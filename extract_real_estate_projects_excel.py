import asyncio
from playwright.async_api import async_playwright
import openpyxl
from bs4 import BeautifulSoup
import re
from docx import Document
import os
import time
import pickle
import ast
import pandas

provinces = [
    'an_giang','ba_ria_vung_tau','bac_kan','bac_lieu','bac_ninh','ben_tre','binh_dinh','binh_phuoc','binh_thuan','ca_mau','can_tho','cao_bang','da_nang','dak_lak','dak_nong','dien_bien','dong_nai','dong_thap',
    'gia_lai','ha_giang','ha_nam','ha_tinh','hai_duong','hau_giang','hoa_binh','hung_yen','khanh_hoa','kien_giang','kon_tum','lai_chau','lam_dong','lang_son','lao_cai','long_an','nam_dinh','nghe_an','ninh_binh',
    'ninh_thuan','phu_tho','phu_yen','quang_binh','quang_nam','quang_ngai','quang_tri','soc_trang','son_la','tay_ninh','thai_binh','thai_binh_duong','thai_nguyen','thanh_hoa','thua_thien_hue','tien_giang','tp_hcm',
    'tra_vinh','tuyen_quang','vinh_long','vinh_phuc','yen_bai'
]

provinces_special = [
    'An Giang','Bà Rịa Vũng Tàu','Bắc Kạn','Bạc Liêu','Bắc Ninh','Bến Tre','Bình Định','Bình Phước','Bình Thuận','Cà Mau','Cần Thơ','Cao Bằng','Đà Nẵng','Đắk Lắk','Đắk Nông','Điện Biên','Đồng Nai','Đồng Tháp',
    'Gia Lai','Hà Giang','Hà Nam','Hà Tĩnh','Hải Dương','Hậu Giang','Hòa Bình','Hưng Yên','Khánh Hòa','Kiên Giang','Kon Tum','Lai Châu','Lâm Đồng','Lạng Sơn','Lào Cai','Long An','Nam Định','Nghệ An','Ninh Bình',
    'Ninh Thuận','Phú Thọ','Phú Yên','Quảng Bình','Quảng Nam','Quảng Ngãi','Quảng Trị','Sóc Trăng','Sơn La','Tây Ninh','Thái Bình','Bình Dương','Thái Nguyên','Thanh Hóa','Thừa Thiên Huế','Tiền Giang',
    'Thành phố Hồ Chí Minh','Trà Vinh','Tuyên Quang','Vĩnh Long','Vĩnh Phúc','Yên Bái'
]

for idx, obj in enumerate(provinces):
    name = obj
    
    parts = name.split('_')
    capitalized_parts = [part.capitalize() for part in parts]
    
    capitalized_name = '_'.join(capitalized_parts)
    
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active

    sheet.cell(row=1,column=1).value = 'Tỉnh'
    sheet.cell(row=1,column=2).value = 'Tên dự án'
    sheet.cell(row=1,column=3).value = 'Loại dự án'
    sheet.cell(row=1,column=4).value = 'Diện tích xây dựng'
    sheet.cell(row=1,column=5).value = 'Diện tích đất'
    sheet.cell(row=1,column=6).value = 'Địa điểm'
    sheet.cell(row=1,column=7).value = 'Loại nguồn vốn'
    sheet.cell(row=1,column=8).value = 'Giá trị'
    sheet.cell(row=1,column=9).value = 'Tình trạng dự án'
    sheet.cell(row=1,column=10).value = 'Ngày bắt đầu'
    sheet.cell(row=1,column=11).value = 'Ngày kết thúc'
    sheet.cell(row=1,column=12).value = 'Vốn đầu tư'
    sheet.cell(row=1,column=13).value = 'Khởi công'
    sheet.cell(row=1,column=14).value = 'Tình trạng'
    sheet.cell(row=1,column=15).value = 'Page source'
    
    df_original = pandas.read_excel(f'C:\\Users\\phams\\Downloads\\du an\\du lieu goc\\{capitalized_name}\\{obj}_projects_links.xlsx')
    
    def extract_between_keywords(big_string, keyword1, keyword2):
        # Find the starting position of keyword1
        start_pos = big_string.find(keyword1)
        
        # If keyword1 is not found, return an empty string
        if start_pos == -1:
            return ""
        
        # Adjust start_pos to the end of keyword1
        start_pos += len(keyword1)
        
        # Find the starting position of keyword2, starting from the end of keyword1
        end_pos = big_string.find(keyword2, start_pos)
        
        # If keyword2 is not found, return an empty string
        if end_pos == -1:
            return ""
        
        # Extract and return the substring between keyword1 and keyword2
        result_str = str(big_string[start_pos:end_pos])
        
        result_str.replace(':','')
        result_str.replace('"','')
        result_str.strip()
        
        return str(result_str)
    
    for i in range(df_original.shape[0]):
        with open(f"C:\\Users\\phams\\Downloads\\du an\\du lieu goc\\{capitalized_name}\\page sources\\page_source_{df_original['Page source'].iloc[i]}.html", 'r', encoding='utf-8') as file:
            html_content = file.read()

        soup = BeautifulSoup(html_content, 'html.parser')

        title = ''
        detail_article = soup.find('article', class_='detail')
        if detail_article:
            h1_tag = detail_article.find('h1')
            if h1_tag:
                title = h1_tag.get_text(separator='\n').strip()

        article_str = ''
        for tag in soup.find_all('article'):
            string = str(tag.get_text(separator='$').strip())
            arr = string.split('$')
            filtered_list = [item.strip() for item in arr if item != '' and item != ' ' and item.strip() != ':' and item.strip() != '']
            article_str = ' '.join(filtered_list)
        
        sheet.cell(row=i+2,column=1).value = provinces_special[idx]
        sheet.cell(row=i+2,column=2).value = title
        sheet.cell(row=i+2,column=15).value = df_original['Page source'].iloc[i]
        
        sheet.cell(row=i+2,column=3).value = extract_between_keywords(article_str,'Loại dự án ','Vị trí dự án ')
        
        sheet.cell(row=i+2,column=4).value = extract_between_keywords(article_str,'Diện tích xây dựng ','Diện tích đất ')
        
        sheet.cell(row=i+2,column=5).value = extract_between_keywords(article_str,'Diện tích đất ','Tình trạng dự án ')
        
        sheet.cell(row=i+2,column=6).value = extract_between_keywords(article_str,'Vị trí dự án ','Loại nguồn vốn ')
        
        sheet.cell(row=i+2,column=7).value = extract_between_keywords(article_str,'Loại nguồn vốn ','Giá trị ')
        
        sheet.cell(row=i+2,column=8).value = extract_between_keywords(article_str,'Giá trị ','Diện tích xây dựng ')
        
        sheet.cell(row=i+2,column=9).value = extract_between_keywords(article_str,'Tình trạng dự án ','Ngày bắt đầu dự án ')
        
        sheet.cell(row=i+2,column=10).value = extract_between_keywords(article_str,'Ngày bắt đầu dự án ','Ngày kết thúc dự án ')
        
        sheet.cell(row=i+2,column=11).value = extract_between_keywords(article_str,'Ngày kết thúc dự án ','Thông tin dự án ')
        
        sheet.cell(row=i+2,column=12).value = extract_between_keywords(article_str,'Vốn đầu tư ','Tình trạng dự án:')
        
        sheet.cell(row=i+2,column=13).value = extract_between_keywords(article_str,'Khởi công:',' *')
        
        sheet.cell(row=i+2,column=14).value = extract_between_keywords(article_str,'* Tình trạng:','CÁC ĐƠN VỊ THAM GIA')
    
    # df = pandas.read_excel(f'C:\\Users\\phams\\Downloads\\du an\\du lieu goc\\{capitalized_name}\\{obj}_check_data.xlsx')
    
    # for i in range(df.shape[0]):
    #     sheet.cell(row=i+2,column=1).value = provinces_special[idx]
    #     sheet.cell(row=i+2,column=14).value = df['Page source'].iloc[i]
    #     sheet.cell(row=i+2,column=2).value = df['Project name'].iloc[i]
        
    #     sheet.cell(row=i+2,column=15).value = str(df['Project name'].iloc[i]) + '.docx'
        
    #     data = str(df['<article>'].iloc[i]).split('&')
        
    #     for x in range(len(data)-3):
    #         if data[x] == 'Loại dự án':
    #             if data[x+1] != 'Vị trí dự án':
    #                 sheet.cell(row=i+2,column=3).value = data[x+1]
    #         if data[x] == 'Vị trí dự án':
    #             if 'Địa bàn' not in data[x+1]:
    #                 string = str(data[x+1]).replace(': ','')
    #                 string = string.replace('Địa điểm: ','')
    #                 string = string.replace('*','')
    #                 string = string.replace('"','')
    #                 sheet.cell(row=i+2,column=5).value = string.strip()
    #         if data[x] == 'Loại nguồn vốn':
    #             if data[x+1] != 'Giá trị':
    #                 sheet.cell(row=i+2,column=6).value = str(data[x+1])
    #         if data[x] == 'Giá trị':
    #             if data[x+1] != 'Diện tích xây dựng' and ',' not in data[x+1]:
    #                 sheet.cell(row=i+2,column=7).value = data[x+1]
    #         if data[x] == 'Tình trạng dự án':
    #             if data[x+1] != 'Ngày bắt đầu dự án':
    #                 sheet.cell(row=i+2,column=8).value = data[x+1]
    #         if data[x] == 'Ngày bắt đầu dự án':
    #             if data[x+1] != 'Ngày kết thúc dự án':
    #                 sheet.cell(row=i+2,column=9).value = data[x+1]
    #         if data[x] == 'Ngày kết thúc dự án':
    #             if data[x+1] != 'Thông tin dự án':
    #                 sheet.cell(row=i+2,column=10).value = data[x+1]
    #         if data[x] == '* Vốn đầu tư':
    #             string = ''
    #             for y in range(x+1,x+3):
    #                 if 'Tình trạng dự án' not in data[y] and 'Triển khai thi công' not in data[y]:
    #                     string = string + data[y] + ' '
    #             string = string.replace('Tình trạng dự án','')
    #             string = string.replace('*','')
    #             string = string.replace(':','')
    #             string = string.replace('"','')
    #             sheet.cell(row=i+2,column=11).value = string.strip()
    #         if data[x] == '* Khởi công:':
    #             string = ''
    #             for y in range(x+1,x+4):
    #                 if 'Diện tích' in data[y] or 'Địa điểm' in data[y] or 'Địa Điểm' in data[y]:
    #                     break
    #                 else:
    #                     string = string + data[y] + ' '
    #             sheet.cell(row=i+2,column=12).value = string.strip()
    #         if data[x] == '* Tình trạng:':
    #             string = ''
    #             for y in range(x+1,x+3):
    #                 if 'CÁC ĐƠN VỊ THAM GIA' in data[y]:
    #                     break
    #                 elif 'Địa điểm' not in data[y] and 'Diện tích' not in data[y]:
    #                     string = string + data[y] + ' '
    #             string = string.replace('Chủ Đầu Tư:','')
    #             string = string.replace('CÁC ĐƠN VỊ THAM GIA','')
    #             string = string.replace('*','')
    #             string = string.replace('"','')
    #             sheet.cell(row=i+2,column=13).value = string.strip()
    #         if data[x] == '* Diện tích':
    #             string = ''
    #             for y in range(x+1,x+4):
    #                 if 'Địa điểm' not in data[x+1]:
    #                     string = string + data[y]
    #             string = string.replace('Tình trạng dự án','')
    #             string = string.replace('*','')
    #             string = string.replace(':','')
    #             string = string.replace('"','')
    #             sheet.cell(row=i+2,column=4).value = string.strip()
    
    spreadsheet.save(f'C:\\Users\\phams\\Downloads\\du an\\du lieu goc\\{capitalized_name}\\{obj}_projects_data.xlsx')
    print(f'Extracted data for {capitalized_name}')