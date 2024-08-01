import asyncio
from playwright.async_api import async_playwright
import json
from datetime import datetime, date, timedelta, timezone
import time
import logging
import signal
import os
import requests
import openpyxl
from openpyxl.styles import Alignment

# --- Configure logging
logging.basicConfig(
    filename="LOG_nghi_dinh.txt",  # --- Replace with your log file path
    level=logging.INFO,  # --- You can set this to DEBUG, INFO, WARNING, ERROR, CRITICAL
    format='%(asctime)s - %(levelname)s - %(message)s',
)

# --- Configure process termination
def signal_handler(sig, frame):
    logging.info("Interrupt received, shutting down...")
    for task in asyncio.all_tasks():
        task.cancel()
    asyncio.get_event_loop().stop()

# --- GHTK chat API
def ghtk_chat(file_path, message):
    headers = {
        'Authorization': 'Bearer eyJhbGciOiJFUzI1NiIsImtpZCI6IjAxRUJRQzVTMlNDVk0wQ0E4TjFKRkpGR0ZHXzE2MDM2MDcyMDAiLCJ0eXAiOiJKV1QifQ.eyJhdWQiOiJhdXRoIiwiZXhwIjoxNzIxNjM1ODA1LCJqdGkiOiIwMUozQ01EQjBWQ0JOTUpDNjNSOTMzRTVDQiIsImlhdCI6MTcyMTYzMjIwNSwiaXNzIjoiaHR0cHM6Ly9hdXRoLmdodGtsYWIuY29tIiwic3ViIjoiMDFHSDg1WDBONTM0UTlaUVRFWlY2WUVFUzkiLCJzY3AiOlsib3BlbmlkIl0sInNpZCI6InV4WHJwaWZyMEtib1VtMUpGREdBbTdWSFh2MzBPQmJpIiwiY2xpZW50X2lkIjoiMDFFQlFDNVMyU0NWTTBDQThOMUpGSkZHRkciLCJ0eXBlIjoib2F1dGgifQ.KwKMOoYsn5i2Q-n1um7JfG_3kzk2vUCzh-_iYFeaQwcPt8zWOTvXZt_vtxh6kmW6F7lz4jDIiVcLOSZwRZn4CQ',
    }

    files = {
        'channel_id': (None, '6997756873994439959'),
        'attachment': open(file_path, 'rb'),
        'msg_type': (None, 'text'),
        'text': (None, message),
    }

    response = requests.post('https://chat.ghtklab.com/api/v3/messages', headers=headers, files=files)
    print(response)

# -- Save extracted data to excel
def save_to_excel(data):
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active

    sheet.cell(row=1,column=1).value = 'Num.'
    sheet.cell(row=1,column=2).value = 'Nghi dinh so'
    sheet.cell(row=1,column=3).value = 'Ngay ban hanh'
    sheet.cell(row=1,column=4).value = 'Ngay hieu luc'
    sheet.cell(row=1,column=5).value = 'Noi dung'
    sheet.cell(row=1,column=6).value = 'Status'

    # Define the column widths
    column_widths = {'A': 8, 'B': 25, 'C': 20, 'D': 20, 'E': 100}

    # Set the column widths
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    if len(data['children']) != 0:
        decrees = data['children'][-2]['children']
        
        total = int(len(decrees) / 4)
        
        for i in range(total):
            sheet.cell(row=i+2,column=1).value = i+1
            for x in range(4):
                sheet.cell(row=i+2,column=x+2).value = decrees[x + (4 * i)]['name']
                if x+2 == 5:
                    sheet.cell(row=i+2,column=x+2).alignment = Alignment(wrapText=True)

    spreadsheet.save('nghi_dinh.xlsx')

if __name__ == "__main__":
    
    signal.signal(signal.SIGINT, signal_handler)
    
    last_checked = datetime.now(tz=timezone.utc)
    reset_time = (datetime.now(tz=timezone.utc) + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    
    try:
        checkpoint = []
        
        # --- Loop the process
        while True:
            today = date.today()
            formatted_today = today.strftime("%d/%m/%Y")
            
            # # --- For testing
            # formatted_today = '01/03/2024'
            
            now = datetime.now(tz=timezone.utc)
            
            if now >= reset_time:
                checkpoint = []
                reset_time = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)

            async def get_accessibility_tree():
                async with async_playwright() as p:
                    # --- Start new Firefox instance
                    # --- user_data_dir points to the Firefox user profile with log in info for GHTK internal chat system
                    # --- To avoid conflicting with any running Firefox instance, use a different user profile containing the same log in info
                    # profile_path = os.getenv('FIREFOX_PROFILE_PATH', 'default/path')
                    browser = await p.firefox.launch_persistent_context(user_data_dir="C:\\Users\\phams\\AppData\\Local\\Mozilla\\Firefox\\Profiles\\bd5r7ma0.default-release-temp", headless=True)
                    
                    # # --- Start new Chrome instance
                    # browser = await p.chromium.launch(headless=True)
                    
                    page = await browser.new_page()
                    
                    await page.goto('https://danhmuchanhchinh.gso.gov.vn/NghiDinh.aspx')
                    
                    # --- Get the accessibility tree (JSON file)
                    accessibility_snapshot = await page.accessibility.snapshot()
                    
                    # --- Save the extracted JSON file to disk (Windows)
                    json_path = 'nghi_dinh.json'
                    
                    with open(json_path, 'w') as f:
                        json.dump(accessibility_snapshot, f, indent=2)

                    # Attempt to open the file with 'utf-8' encoding
                    try:
                        with open(json_path, 'r', encoding='utf-8') as file:
                            data = json.load(file)
                    except UnicodeDecodeError:
                        # If 'utf-8' encoding fails, try 'iso-8859-1'
                        with open(json_path, 'r', encoding='iso-8859-1') as file:
                            data = json.load(file)
                    
                    new_decrees = []
                    count = 0
                    
                    # --- Check for the "children" node in the JSON file, which contains info about decrees - Firefox
                    if len(data['children']) != 0:
                        decrees = data['children'][-2]['children']
                        total = int(len(decrees) / 4)
                        
                        for i in range(total):
                            if datetime.strptime(decrees[1 + (4 * i)]['name'], "%d/%m/%Y").date() >= datetime.strptime(formatted_today, "%d/%m/%Y").date() and (decrees[0 + (4 * i)]['name'],formatted_today) not in checkpoint:
                                new = {
                                    'nghi_dinh_so': decrees[0 + (4 * i)]['name'],
                                    'ngay_ban_hanh': decrees[1 + (4 * i)]['name'],
                                    'ngay_hieu_luc': decrees[2 + (4 * i)]['name'],
                                    'noi_dung': decrees[3 + (4 * i)]['name']
                                }
                                new_decrees.append(new)
                                checkpoint.append((decrees[0 + (4 * i)]['name'],formatted_today))
                                count += 1

                    # # --- Check for the "children" node in the JSON file, which contains info about decrees - Chrome
                    # if len(data['children']) != 0:
                    #     decrees = data['children']
                    #     total = int((len(decrees) - 20) / 4)
                        
                    #     for i in range(total):
                    #         # --- Check for new decree(s) based on current date and see if that decree(s) has been checked or not,
                    #         # --- to prevent duplicating within the same day
                    #         if datetime.strptime(decrees[17 + (4 * i)]['name'], "%d/%m/%Y").date() >= datetime.strptime(formatted_today, "%d/%m/%Y").date() and (decrees[16 + (4 * i)]['name'],formatted_today) not in checkpoint:
                    #             new = {
                    #                 'nghi_dinh_so': decrees[16 + (4 * i)]['name'],
                    #                 'ngay_ban_hanh': decrees[17 + (4 * i)]['name'],
                    #                 'ngay_hieu_luc': decrees[18 + (4 * i)]['name'],
                    #                 'noi_dung': decrees[19 + (4 * i)]['name']
                    #             }
                    #             new_decrees.append(new)
                    #             checkpoint.append((decrees[0 + (4 * i)]['name'],formatted_today))
                    #             count += 1

                    message = f'Automated message: Ngày {formatted_today} có {count} nghị định mới'
                    
                    # --- Navigate to the desired GHTK group chat using URL
                    await page.goto('https://ghtk.me/channel/internal/recent/5410781654964860139')

                    # --- Prepare to log the result for new decree(s)
                    if count > 0:
                        new_decrees_info = []
                        
                        save_to_excel(data)
                        
                        # --- Set up new decree(s) info
                        for data in new_decrees:
                            template = '''\nNghị định số: {nghi_dinh_so}, \nNgày ban hành: {ngay_ban_hanh}, \nNgày hiệu lực: {ngay_hieu_luc}, \nNội dung: {noi_dung}'''
                            result = f'{template.format(**data)}'
                            new_decrees_info.append(result)

                        new_decrees_info_merge = '\n'.join(new_decrees_info)

                        message = message + new_decrees_info_merge + ' '
                        
                        # # Wait for the file upload button to be available
                        # await page.wait_for_selector('[class="footer-view__rep"]')
                        # await page.click('[class="footer-view__rep"]')

                        # # Set the file to be uploaded using the hidden input element
                        # file_input_selector = 'input[type="file"]'  # Replace with the actual selector for the file input
                        # await page.set_input_files(file_input_selector, "C:\\Users\\Public\\Documents\\vscode_python\\nghi_dinh.xlsx")
                        
                        # --- Wait for the chat box to load
                        await page.wait_for_selector('div[contenteditable="true"]', timeout=50000)
                        
                        # --- Find the chat box and type a message
                        chat_box = await page.query_selector('div[contenteditable="true"]')
                        for char in message:
                            if char == '\n':
                                await chat_box.press('Shift+Enter')
                            else:
                                await chat_box.type(char, delay=50)
                        
                        # --- Press Enter to send the message
                        async def click_send_button():
                            await page.wait_for_selector('[class="footer-view__rep"]', timeout=20000)
                            send_button = await page.query_selector('[class="footer-view__rep"]')
                            if send_button:
                                await send_button.click()
                                return True
                            else:
                                return False
                                
                        while not await click_send_button():
                            await asyncio.sleep(2)
                        
                        logging.info(message)
                    else:
                        message = message + ' '
                        
                        # --- Wait for the chat box to load
                        await page.wait_for_selector('div[contenteditable="true"]', timeout=50000)
                        
                        # --- Find the chat box and type a message
                        chat_box = await page.query_selector('div[contenteditable="true"]')
                        for char in message:
                            if char == '\n':
                                await chat_box.press('Shift+Enter')
                            else:
                                await chat_box.type(char, delay=50)
                        
                        # --- Press Enter to send the message
                        async def click_send_button():
                            await page.wait_for_selector('[class="footer-view__rep"]', timeout=20000)
                            send_button = await page.query_selector('[class="footer-view__rep"]')
                            if send_button:
                                await send_button.click()
                                return True
                            else:
                                return False
                                
                        while not await click_send_button():
                            await asyncio.sleep(2)
                        
                        # --- Log the result for no new decree
                        logging.info(f': No new decree')
                    
                    # --- Close the browser
                    await browser.close()

            # --- Run the process
            
            logging.info(f': Starting browser to check for new decrees...')
            
            asyncio.get_event_loop().run_until_complete(get_accessibility_tree())
            
            # --- Set the process to sleep for a set amount of time
            interval = 18000 # seconds
            
            logging.info(f': Going into sleep mode for {interval / 3600} hours...')
            
            time.sleep(interval)
            
    except Exception as e:
        # --- Log any detected error
        logging.critical(f"Fatal error: {e}")
    finally:
        # --- Log the process termination
        logging.info("Monitoring stopped")