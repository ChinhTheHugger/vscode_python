import openpyxl
import pandas
from collections import Counter
from difflib import SequenceMatcher
from collections import OrderedDict
import time
import numpy
import igraph

import sys

pathTest = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book1.XLSX"
path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.17 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V24 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

wb_test = openpyxl.load_workbook(pathTest)
sh_test = wb_test.active

start = time.time()
print("Start time = " + time.strftime("%H:%M:%S", time.gmtime(start)))
time.sleep(5)

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=7).value = sheet_obj.cell(row=x,column=6).value

def birthdayGraph(startNum,endNum):
    bArray = []
    for x in range(startNum,endNum):
        if str(sheet_obj.cell(row=x,column=4).value) not in bArray:
            bArray.append(str(sheet_obj.cell(row=x,column=4).value))
        if str(sheet_obj.cell(row=x,column=5).value) not in bArray:
            bArray.append(str(sheet_obj.cell(row=x,column=5).value))
    bArray = sorted(bArray)
    for a in range(startNum,endNum):
        sheet_obj.cell(row=a,column=6).value = bArray.index(str(sheet_obj.cell(row=a,column=4).value))
        sheet_obj.cell(row=a,column=7).value = bArray.index(str(sheet_obj.cell(row=a,column=5).value))
    bGraph = igraph.Graph(n=len(bArray))
    bGraph.vs['name'] = bArray
    for x in range(startNum,endNum):
        bGraph.add_edges([(sheet_obj.cell(row=x,column=6).value,sheet_obj.cell(row=x,column=7).value)])
    for s in bGraph.components().subgraphs():
        bArr = []
        for ele in s.vs['name']:
            if ";" not in ele:
                bArr.append(ele)
        for x in range(startNum,endNum):
            if str(sheet_obj.cell(row=x,column=5).value) in bArr:
                sheet_obj.cell(row=x,column=8).value = ';'.join(sorted(bArr))

arr = [2]
for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,sheet_obj.max_row+2):
        if sheet_obj.cell(row=i,column=3).value != sheet_obj.cell(row=x,column=3).value:
            break
    if i not in arr:
        arr.append(i)
    x=i

# print(arr)
print(len(arr))

newArr = numpy.array(arr)
for x in range(0,len(newArr)-1):
    if newArr[x+1]-newArr[x]==1:
        sheet_obj.cell(row=newArr[x],column=8).value = str(sheet_obj.cell(row=newArr[x],column=5).value)
    else:
        for i in range(newArr[x],newArr[x+1]):
            birthdayGraph(newArr[x],newArr[x+1])

wb_obj.save(path)

end = time.time()

print("Run time = " + time.strftime("%H:%M:%S", time.gmtime(end-start)))
print("End time = " + time.strftime("%H:%M:%S", time.gmtime(end)))