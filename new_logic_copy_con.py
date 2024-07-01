import openpyxl
import pandas
from collections import Counter
from difflib import SequenceMatcher
from collections import OrderedDict
import time
import numpy

pathTest = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book.XLSX"
path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.07 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()

# Continuation from new_logic_copy.py

# New idea
# Compare string similarity - WRONG
# If the result is larger than 0, then we can merge data - WRONG
# String similarity compare EACH character in the strings, so techincally, ALL strings share zimilarity larger than 0%

# Lets try pandas solution
# Doesnt work

# Lets try list comprehension
# List comprehension seems to work



# test

# s = SequenceMatcher(None, str(sheet_obj.cell(row=508, column=6).value), str(sheet_obj.cell(row=509, column=6).value))
# print(s.ratio())

# a = pandas.Series(str(sheet_obj.cell(row=508, column=6).value))
# b = pandas.Series(str(sheet_obj.cell(row=509, column=6).value))
# matching_percentage = (a==b).mean()
# print(matching_percentage)

# arrOne = numpy.array(str(sheet_obj.cell(row=508, column=6).value).split(';'))
# strOne = str(sheet_obj.cell(row=507, column=6).value)
# print (sum([1 for i in range(len(arrOne)) if arrOne[i] in strOne]))



# House group

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpArr.sort()
#     tmpStrN = ';'.join(tmpArr)
#     return tmpStrN

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=7).value=sheet_obj.cell(row=x,column=5).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=7).value=sheet_obj.cell(row=a,column=5).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=7).value))
#             for b in range(x,i+1):
#                 strOne = str(sheet_obj.cell(row=a,column=6).value)
#                 arrTwo = numpy.array(str(sheet_obj.cell(row=b,column=6).value).split(';'))
#                 if sum([1 for i in range(len(arrTwo)) if arrTwo[i] in strOne]) > 0:
#                     houseArr.append(str(sheet_obj.cell(row=b,column=5).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=7).value=houseStr
#     return          

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

def cleanPhone(arr):
    tmpStr=';'.join(arr)
    tmpArr=tmpStr.split(';')
    tmpArr=list(dict.fromkeys(tmpArr))
    tmpArr.sort()
    tmpStrN = ';'.join(tmpArr)
    return tmpStrN

def phoneGroup(x,i):
    if i-x==0:
        sheet_obj.cell(row=x,column=8).value=sheet_obj.cell(row=x,column=6).value
    else:
        for a in range(x,i+1):
            sheet_obj.cell(row=a,column=8).value=sheet_obj.cell(row=a,column=6).value
            phoneArr=[]
            phoneArr.append(str(sheet_obj.cell(row=a,column=8).value))
            for b in range(x,i+1):
                strOne = str(sheet_obj.cell(row=a,column=7).value)
                arrTwo = numpy.array(str(sheet_obj.cell(row=b,column=7).value).split(';'))
                if sum([1 for i in range(len(arrTwo)) if arrTwo[i] in strOne]) > 0:
                    phoneArr.append(str(sheet_obj.cell(row=b,column=6).value))
            phoneStr=cleanPhone(phoneArr)
            sheet_obj.cell(row=a,column=8).value=phoneStr
    return          

arr = [2]
for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,sheet_obj.max_row+2):
        if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
            break
    if i not in arr:
        arr.append(i)
    x=i

print(arr)
print(len(arr))

newArr = numpy.array(arr)
for x in range(0,len(arr)-1):
    phoneGroup(newArr[x],newArr[x+1]-1)



wb_obj.save(path)

end = time.time()

print("Run time = " + time.strftime("%H:%M:%S", time.gmtime(end-start)))