import openpyxl
from collections import Counter
import time

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.09 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V21 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()

# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+1):
#         if sheet_obj.cell(row=i,column=3).value != sheet_obj.cell(row=x,column=3).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=6).value == sheet_obj.cell(row=x,column=6).value:
#                 sheet_obj.cell(row=i,column=1).value = None
#     x=i

# ***

# * Group house by phone, merge and clean

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row = x, column = 8).value = sheet_obj.cell(row = x, column = 7).value
#     for i in range(x+1,x+5000):
#         if sheet_obj.cell(row = i, column = 2).value == sheet_obj.cell(row = x, column = 2).value and sheet_obj.cell(row = i, column = 4).value == sheet_obj.cell(row = x, column = 4).value:
#             sheet_obj.cell(row = x, column = 8).value = sheet_obj.cell(row = x, column = 8).value + ";" + sheet_obj.cell(row = i, column = 7).value
#             sheet_obj.cell(row = i, column = 3).value = None
#             continue
#         else:
#             break
#     x = i

# for x in range(2,sheet_obj.max_row+1):
#     if sheet_obj.cell(row=x,column=3).value != None:
#         for i in range(x+1,sheet_obj.max_row+1):
#             if sheet_obj.cell(row=i,column=3).value == None:
#                 sheet_obj.cell(row=i,column=8).value = sheet_obj.cell(row=x,column=8).value
#     x = i

# for x in range(2,sheet_obj.max_row+1):
#     houseString = sheet_obj.cell(row = x, column = 8).value
#     houseArrayToSort = houseString.split(';')
#     houseArrayToSort = list(set(houseArrayToSort))
#     houseListToStr = ';'.join(map(str,houseArrayToSort))
#     sheet_obj.cell(row = x, column = 8).value = houseListToStr

# * Group phone by house, merge and clean

for x in range(2,sheet_obj.max_row+1):
    sheet_obj.cell(row = x, column = 9).number_format = '@' 
    sheet_obj.cell(row = x, column = 9).value = sheet_obj.cell(row = x, column = 4).value
    for i in range(x+1,x+5000):
        if sheet_obj.cell(row = i, column = 2).value == sheet_obj.cell(row = x, column = 2).value and sheet_obj.cell(row = i, column = 7).value == sheet_obj.cell(row = x, column = 7).value:
            temp = sheet_obj.cell(row = x, column = 9).value + ";" + sheet_obj.cell(row = i, column = 4).value
            sheet_obj.cell(row = x, column = 9).value = temp
            sheet_obj.cell(row = i, column = 3).value = None
            continue
        else:
            break
    x = i

for x in range(2,sheet_obj.max_row+1):
    if sheet_obj.cell(row=x,column=3).value != None:
        for i in range(x+1,x+5000):
            if sheet_obj.cell(row=i,column=3).value == None:
                sheet_obj.cell(row=i,column=9).value = sheet_obj.cell(row=x,column=9).value
    x = i

for x in range(2,sheet_obj.max_row+1):
    phoneString = sheet_obj.cell(row = x, column = 9).value
    phoneArrayToSort = phoneString.split(';')
    phoneArrayToSort = list(set(phoneArrayToSort))
    phoneListToStr = ';'.join(map(str,phoneArrayToSort))
    sheet_obj.cell(row = x, column = 9).value = phoneListToStr

wb_obj.save(path)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))