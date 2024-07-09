import json
import openpyxl

# spreadsheet = openpyxl.Workbook()
# sheet = spreadsheet.active

spreadsheet = openpyxl.load_workbook("C:\\Users\\phams\\Downloads\\Hung_Yen\\hung_yen_projects_links.xlsx")
sheet = spreadsheet.active

file_path = "C:\\Users\\phams\\Downloads\\hung_yen_page_2.json"

# Attempt to open the file with 'utf-8' encoding
try:
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
except UnicodeDecodeError:
    # If 'utf-8' encoding fails, try 'iso-8859-1'
    with open(file_path, 'r', encoding='iso-8859-1') as file:
        data = json.load(file)
        
links = []

if len(data['children']) != 0:
    for i in range(len(data['children'])):
        if len(data['children'][i]['children']) != 0:
            if data['children'][i]['children'][0]['value'] not in links and data['children'][i]['children'][0]['value'] != '':
                links.append(data['children'][i]['children'][0]['value'])

# print(len(links))

sheet.cell(row=1,column=1).value = 'Page source'
sheet.cell(row=1,column=2).value = 'Links'
sheet.cell(row=1,column=3).value = 'Project name'

for x in range(len(links)):
    sheet.cell(row=x+102,column=1).value = x+101
    sheet.cell(row=x+102,column=2).value = links[x]
    
sheet_file = "C:\\Users\\phams\\Downloads\\Hung_Yen\\hung_yen_projects_links.xlsx"
spreadsheet.save(sheet_file)

