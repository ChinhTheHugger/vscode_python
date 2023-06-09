import openpyxl
from collections import Counter
import time
import numpy

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.21 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V22 - processed.XLSX"
# path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.18 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()
print("Start time = " + time.strftime("%H:%M:%S", time.gmtime(start)))
time.sleep(5)

# ***

# * Count general phone and house

# for x in range(2,sheet_obj.max_row+1):
#     # #phone
#     phoneCount = ""
#     phoneCount = sheet_obj.cell(row = x, column = 15).value
#     sheet_obj.cell(row = x, column = 14).value = phoneCount.count(';') + 1
#     #house
#     houseCount = ""
#     houseCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 13).value = houseCount.count(';')+  1

# ***

# * Count house in each area

# for x in range(2,sheet_obj.max_row+1):
#     #VRSH
#     areaCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 5).value = areaCount.count('VRSH')
#     #VOP1
#     areaCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 6).value = areaCount.count('VOP1')
#     #VOP2
#     areaCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 7).value = areaCount.count('VOP2')
#     #ECP
#     areaCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 8).value = areaCount.count('ECP')
#     #STL
#     areaCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 9).value = areaCount.count('STL')
#     #VGP
#     areaCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 10).value = areaCount.count('VGP')
#     #VGV
#     areaCount = sheet_obj.cell(row = x, column = 4).value
#     sheet_obj.cell(row = x, column = 11).value = areaCount.count('VGV')

# ***

# * Count total area

# for x in range(2,sheet_obj.max_row+1):
#     count = 0
#     for i in range(5,11):
#         if sheet_obj.cell(row = x, column = i).value > 0:
#             count += 1
#     sheet_obj.cell(row = x, column = 12).value = count

# Sort house

# for x in range(2,sheet_obj.max_row+1):
#     arr=str(sheet_obj.cell(row=x,column=4).value).split(';')
#     arr.sort()
#     newStr=';'.join(arr)
#     sheet_obj.cell(row=x,column=4).value=newStr

# Sort phone

# for x in range(2,sheet_obj.max_row+1):
#     arr=str(sheet_obj.cell(row=x,column=18).value).split(';')
#     arr.sort()
#     newStr=';'.join(arr)
#     sheet_obj.cell(row=x,column=18).value=newStr



wb_obj.save(path)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))
print("End time = " + time.strftime("%H:%M:%S", time.gmtime(end)))