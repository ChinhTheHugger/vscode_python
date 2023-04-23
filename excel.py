import openpyxl
from collections import Counter
from difflib import SequenceMatcher
from collections import OrderedDict
import time
import numpy

import sys

path = "F:\\Book1.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_base = wb_obj.worksheets[0]
sheet_area_1 = wb_obj.worksheets[1]
sheet_area_2 = wb_obj.worksheets[2]

start = time.time()
print("Start time = " + time.strftime("%H:%M:%S", time.gmtime(start)))
time.sleep(5)

# hArr = []
# aArr = []
# for x in range(2,sheet_area_1.max_row+1):
#     if sheet_area_1.cell(row=x,column=1).value not in hArr:
#         hArr.append(str(sheet_area_1.cell(row=x,column=1).value))
#         aArr.append(str(sheet_area_1.cell(row=x,column=2).value))
# nhArr = numpy.array(hArr)
# naArr = numpy.array(aArr)
# for i in range(0,len(nhArr)):
#     sheet_area_2.cell(row=i+2,column=1).value = nhArr[i]
#     sheet_area_2.cell(row=i+2,column=2).value = naArr[i]

for x in range(2,sheet_base.max_row+1):
    count = 0
    areaSum = 0
    for i in range(2,sheet_area_2.max_row+1):
        if str(sheet_area_2.cell(row=i,column=1).value) in str(sheet_base.cell(row=x,column=3).value):
            count = count + 1
            areaSum = areaSum + float(sheet_area_2.cell(row=i,column=2).value)
    sheet_base.cell(row=x,column=6).value = areaSum
    sheet_base.cell(row=x,column=7).value = count

wb_obj.save(path)

end = time.time()

print("Run time = " + time.strftime("%H:%M:%S", time.gmtime(end-start)))