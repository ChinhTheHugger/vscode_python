import openpyxl
from collections import Counter
import time
import numpy

pathTest = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book.XLSX"
path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.06 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()

# * First step *

# House group

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=5).value=sheet_obj.cell(row=x,column=4).value
#     else:
#         for a in range(x,i+1):
#             houseArrTemp=[]
#             houseArrTemp.append(sheet_obj.cell(row=a,column=4).value)
#             for b in range(x,i+1):
#                 if sheet_obj.cell(row=b,column=3).value==sheet_obj.cell(row=a,column=3).value and sheet_obj.cell(row=b,column=4).value not in houseArrTemp:
#                     houseArrTemp.append(sheet_obj.cell(row=b,column=4).value)
#             # houseArr=list(dict.fromkeys(houseArrTemp))
#             houseStr=';'.join(map(str,houseArrTemp))
#             sheet_obj.cell(row=a,column=5).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group - this one will be used to group people sharing phone later

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=6).value=sheet_obj.cell(row=x,column=3).value
#     else:
#         for a in range(x,i+1):
#             phoneArrTemp=[]
#             phoneArrTemp.append(sheet_obj.cell(row=a,column=3).value)
#             for b in range(x,i+1):
#                 if sheet_obj.cell(row=b,column=4).value==sheet_obj.cell(row=a,column=4).value and sheet_obj.cell(row=b,column=3).value not in phoneArrTemp:
#                     phoneArrTemp.append(sheet_obj.cell(row=b,column=3).value)
#             # houseArr=list(dict.fromkeys(houseArrTemp))
#             phoneStr=';'.join(map(str,phoneArrTemp))
#             sheet_obj.cell(row=a,column=6).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     phoneGroup(newArr[x],newArr[x+1]-1)

# * Second step *

# * List all pseudo column C into different cell in a new file

# wb_test=openpyxl.load_workbook(pathTest)
# sh_test=wb_test.active

# for x in range(2,sheet_obj.max_row+1):
#     arr=str(sheet_obj.cell(row=x,column=5).value).split(';')
#     sh_test.cell(row=x,column=1).value=len(arr)
#     if len(arr)==1:
#         sh_test.cell(row=x,column=3).value=sheet_obj.cell(row=x,column=5).value
#     else:
#         for i in range(0,len(arr)):
#             sh_test.cell(row=x,column=i+3).value=arr[i]

# wb_test.save(pathTest)

# House group

# wb_test=openpyxl.load_workbook(pathTest)
# sh_test=wb_test.active

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=7).value=sheet_obj.cell(row=x,column=5).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=7).value=sheet_obj.cell(row=a,column=5).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=7).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=3).value) in str(sheet_obj.cell(row=b,column=6).value):
#                     houseArr.append(str(sheet_obj.cell(row=b,column=5).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=7).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

# wb_test=openpyxl.load_workbook(pathTest)
# sh_test=wb_test.active

# def cleanPhone(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=9).value=sheet_obj.cell(row=x,column=6).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=9).value=sheet_obj.cell(row=a,column=6).value
#             phoneArr=[]
#             phoneArr.append(str(sheet_obj.cell(row=a,column=9).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=4).value) in str(sheet_obj.cell(row=b,column=5).value):
#                     phoneArr.append(str(sheet_obj.cell(row=b,column=6).value))
#             phoneStr=cleanPhone(phoneArr)
#             sheet_obj.cell(row=a,column=9).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     phoneGroup(newArr[x],newArr[x+1]-1)

# * Third step, the same as second step logically *

# House group

# wb_test=openpyxl.load_workbook(pathTest)
# sh_test=wb_test.active

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=11).value=sheet_obj.cell(row=x,column=7).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=11).value=sheet_obj.cell(row=a,column=7).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=11).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=3).value) in str(sheet_obj.cell(row=b,column=9).value):
#                     houseArr.append(str(sheet_obj.cell(row=b,column=7).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=11).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

# def cleanPhone(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=13).value=sheet_obj.cell(row=x,column=9).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=13).value=sheet_obj.cell(row=a,column=9).value
#             phoneArr=[]
#             phoneArr.append(str(sheet_obj.cell(row=a,column=9).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=4).value) in str(sheet_obj.cell(row=b,column=7).value):
#                     phoneArr.append(str(sheet_obj.cell(row=b,column=9).value))
#             phoneStr=cleanPhone(phoneArr)
#             sheet_obj.cell(row=a,column=13).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# * after cheching, it seems like three steps may not be enough to get an accurate result
# * or the process strings are too large, and python cant group and remove duplicates properly
# * fourth step, same as third and second

# House group

# wb_test=openpyxl.load_workbook(pathTest)
# sh_test=wb_test.active

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=15).value=sheet_obj.cell(row=x,column=12).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=15).value=sheet_obj.cell(row=a,column=12).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=15).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=3).value) in str(sheet_obj.cell(row=b,column=14).value):
#                     houseArr.append(str(sheet_obj.cell(row=b,column=12).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=15).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

# def cleanPhone(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=17).value=sheet_obj.cell(row=x,column=14).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=17).value=sheet_obj.cell(row=a,column=14).value
#             phoneArr=[]
#             phoneArr.append(str(sheet_obj.cell(row=a,column=17).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=4).value) in str(sheet_obj.cell(row=b,column=12).value):
#                     phoneArr.append(str(sheet_obj.cell(row=b,column=14).value))
#             phoneStr=cleanPhone(phoneArr)
#             sheet_obj.cell(row=a,column=17).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     phoneGroup(newArr[x],newArr[x+1]-1)

# !!!!! *** !!!!!

# * Fifth step, different from previous four steps-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Group house first, same as the first four steps (S1)
# Group phone, but instead of using the house group result from fourth step, use result from S1 above (S2)
# Group house again, the same as above, which means using result from S2
# Maybe group phone once more

# After checking, notice that some combo have very high similarity
# But even after 4 steps, still cant be grouped together
# Which mean even the original data of the combos already have no similarity from the start
# So let's try grouping using the result from the same step

# FIRST BATCH OF HOUSE AND PHONE GROUP-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# House group

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=7).value=sheet_obj.cell(row=x,column=5).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=7).value=sheet_obj.cell(row=a,column=5).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=7).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=3).value) in str(sheet_obj.cell(row=b,column=6).value):
#                     houseArr.append(str(sheet_obj.cell(row=b,column=5).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=7).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

# def cleanPhone(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=8).value=sheet_obj.cell(row=x,column=6).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=8).value=sheet_obj.cell(row=a,column=6).value
#             phoneArr=[]
#             phoneArr.append(str(sheet_obj.cell(row=a,column=8).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=4).value) in str(sheet_obj.cell(row=b,column=7).value):
#                     phoneArr.append(str(sheet_obj.cell(row=b,column=6).value))
#             phoneStr=cleanPhone(phoneArr)
#             sheet_obj.cell(row=a,column=8).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     phoneGroup(newArr[x],newArr[x+1]-1)

# SECOND BATCH OF HOUSE AND PHONE GROUP----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# House group

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=9).value=sheet_obj.cell(row=x,column=7).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=9).value=sheet_obj.cell(row=a,column=7).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=9).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=3).value) in str(sheet_obj.cell(row=b,column=8).value):
#                     houseArr.append(str(sheet_obj.cell(row=b,column=7).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=9).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

# def cleanPhone(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=10).value=sheet_obj.cell(row=x,column=8).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=10).value=sheet_obj.cell(row=a,column=8).value
#             phoneArr=[]
#             phoneArr.append(str(sheet_obj.cell(row=a,column=10).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=4).value) in str(sheet_obj.cell(row=b,column=9).value):
#                     phoneArr.append(str(sheet_obj.cell(row=b,column=8).value))
#             phoneStr=cleanPhone(phoneArr)
#             sheet_obj.cell(row=a,column=10).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     phoneGroup(newArr[x],newArr[x+1]-1)

# THIRD BATCH OF HOUSE AND PHONE GROUP-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# House group

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=11).value=sheet_obj.cell(row=x,column=9).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=11).value=sheet_obj.cell(row=a,column=9).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=11).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=3).value) in str(sheet_obj.cell(row=b,column=10).value):
#                     houseArr.append(str(sheet_obj.cell(row=b,column=9).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=11).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

# def cleanPhone(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=12).value=sheet_obj.cell(row=x,column=10).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=12).value=sheet_obj.cell(row=a,column=10).value
#             phoneArr=[]
#             phoneArr.append(str(sheet_obj.cell(row=a,column=12).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=4).value) in str(sheet_obj.cell(row=b,column=11).value):
#                     phoneArr.append(str(sheet_obj.cell(row=b,column=10).value))
#             phoneStr=cleanPhone(phoneArr)
#             sheet_obj.cell(row=a,column=12).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     phoneGroup(newArr[x],newArr[x+1]-1)

# FOURTH BATCH OF HOUSE AND PHONE GROUP----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# House group

# def cleanHouse(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=15).value=sheet_obj.cell(row=x,column=11).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=15).value=sheet_obj.cell(row=a,column=11).value
#             houseArr=[]
#             houseArr.append(str(sheet_obj.cell(row=a,column=15).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=3).value) in str(sheet_obj.cell(row=b,column=12).value):
#                     houseArr.append(str(sheet_obj.cell(row=b,column=11).value))
#             houseStr=cleanHouse(houseArr)
#             sheet_obj.cell(row=a,column=15).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

# def cleanPhone(arr):
#     tmpStr=';'.join(arr)
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpStr=';'.join(tmpArr)
#     return tmpStr

# def phoneGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=16).value=sheet_obj.cell(row=x,column=12).value
#     else:
#         for a in range(x,i+1):
#             sheet_obj.cell(row=a,column=16).value=sheet_obj.cell(row=a,column=12).value
#             phoneArr=[]
#             phoneArr.append(str(sheet_obj.cell(row=a,column=16).value))
#             for b in range(x,i+1):
#                 if str(sheet_obj.cell(row=a,column=4).value) in str(sheet_obj.cell(row=b,column=15).value):
#                     phoneArr.append(str(sheet_obj.cell(row=b,column=12).value))
#             phoneStr=cleanPhone(phoneArr)
#             sheet_obj.cell(row=a,column=16).value=phoneStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-1):
#     phoneGroup(newArr[x],newArr[x+1]-1)



# Clean and sort house

# for x in range(2,sheet_obj.max_row+1):
#     tmpStr = str(sheet_obj.cell(row=x,column=15).value)
#     tmpArr = tmpStr.split(';')
#     tmpArr.sort()
#     tmpStrN = ';'.join(tmpArr)
#     sheet_obj.cell(row=x,column=17).value = tmpStrN

# Clean and sort phone

# for x in range(2,sheet_obj.max_row+1):
#     tmpStr = str(sheet_obj.cell(row=x,column=16).value)
#     tmpArr = tmpStr.split(';')
#     tmpArr.sort()
#     tmpStrN = ';'.join(tmpArr)
#     sheet_obj.cell(row=x,column=18).value = tmpStrN

# Count house and phone length

for x in range(2, sheet_obj.max_row+1):
    sheet_obj.cell(row=x,column=19).value = str(sheet_obj.cell(row=x,column=17).value).count(';') + 1
    sheet_obj.cell(row=x,column=20).value = str(sheet_obj.cell(row=x,column=18).value).count(';') + 1

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=17).value = "H" + str(sheet_obj.cell(row=x,column=15).value) + " + " + "P" + str(sheet_obj.cell(row=x,column=16).value)



wb_obj.save(path)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))