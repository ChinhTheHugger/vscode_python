import openpyxl
import polyline
from shapely import LineString
import numpy as np
import math



# Google excel sheet
spreadsheet_gg = openpyxl.load_workbook("C:\\Users\\phams\\Downloads\\failed_case.xlsx")

sheet_gg = spreadsheet_gg.active

# result sheet
spreadsheet = openpyxl.Workbook()
sheet = spreadsheet.active

# Haversine distance function - return in kilometer
def haversine(lon1, lat1, lon2, lat2):
    # Radius of the Earth in km
    R = 6371.0

    # Convert decimal degrees to radians
    lat1 = math.radians(lat1)
    lon1 = math.radians(lon1)
    lat2 = math.radians(lat2)
    lon2 = math.radians(lon2)

    # Differences in coordinates
    dlat = lat2 - lat1
    dlon = lon2 - lon1

    # Haversine formula
    a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    # Distance in kilometers
    distance = R * c

    return distance



sheet.cell(row=1,column=1).value = 'Case'

sheet.cell(row=1,column=2).value = 'Start long - original'
sheet.cell(row=1,column=3).value = 'Start lat - original'

sheet.cell(row=1,column=4).value = 'End long - original'
sheet.cell(row=1,column=5).value = 'End lat - original'

sheet.cell(row=1,column=6).value = 'Start long - from GG result'
sheet.cell(row=1,column=7).value = 'Start lat - from GG result'

sheet.cell(row=1,column=8).value = 'End long - from GG result'
sheet.cell(row=1,column=9).value = 'End lat - from GG result'

sheet.cell(row=1,column=10).value = 'Similar start point'
sheet.cell(row=1,column=11).value = 'Similar end point'

sheet.cell(row=1,column=12).value = 'Start lon deviation'
sheet.cell(row=1,column=13).value = 'Start lat deviation'

sheet.cell(row=1,column=14).value = 'End lon deviation'
sheet.cell(row=1,column=15).value = 'End lat deviation'

sheet.cell(row=1,column=16).value = 'Start points haversine distance (m)'
sheet.cell(row=1,column=17).value = 'End points haversine distance (m)'

for i in range(2,sheet_gg.max_row+1):
    sheet.cell(row=i,column=1).value = sheet_gg.cell(row=i,column=1).value
    
    sheet.cell(row=i,column=2).value = sheet_gg.cell(row=i,column=2).value
    sheet.cell(row=i,column=3).value = sheet_gg.cell(row=i,column=3).value
    
    sheet.cell(row=i,column=4).value = sheet_gg.cell(row=i,column=4).value
    sheet.cell(row=i,column=5).value = sheet_gg.cell(row=i,column=5).value
    
    google_encoded = str(sheet_gg.cell(row=i,column=6).value)
    google_decoded = polyline.decode(google_encoded)
    google_line = LineString(google_decoded)
    google_line = LineString([[lon, lat] for lat, lon in google_line.coords])
    
    gg_start_lon, gg_start_lat = google_line.coords[0]
    gg_end_lon, gg_end_lat = google_line.coords[-1]
    
    sheet.cell(row=i,column=6).value = gg_start_lon
    sheet.cell(row=i,column=7).value = gg_start_lat
    
    sheet.cell(row=i,column=8).value = gg_end_lon
    sheet.cell(row=i,column=9).value = gg_end_lat
    
    if sheet_gg.cell(row=i,column=2).value == gg_start_lon and sheet_gg.cell(row=i,column=3).value == gg_start_lat:
        sheet.cell(row=i,column=10).value = 1
        
        sheet.cell(row=i,column=12).value = 0
        sheet.cell(row=i,column=13).value = 0
        
        sheet.cell(row=i,column=16).value = 0
    else:
        sheet.cell(row=i,column=10).value = 0
        
        sheet.cell(row=i,column=12).value = gg_start_lon - sheet_gg.cell(row=i,column=2).value
        sheet.cell(row=i,column=13).value = gg_start_lat - sheet_gg.cell(row=i,column=3).value
        
        sheet.cell(row=i,column=16).value = haversine(sheet_gg.cell(row=i,column=2).value,sheet_gg.cell(row=i,column=3).value,gg_start_lon,gg_start_lat) * 1000
    
    if sheet_gg.cell(row=i,column=4).value == gg_end_lon and sheet_gg.cell(row=i,column=5).value == gg_end_lat:
        sheet.cell(row=i,column=11).value = 1
        
        sheet.cell(row=i,column=14).value = 0
        sheet.cell(row=i,column=15).value = 0
        
        sheet.cell(row=i,column=17).value = 0
    else:
        sheet.cell(row=i,column=11).value = 0
        
        sheet.cell(row=i,column=14).value = gg_end_lon - sheet_gg.cell(row=i,column=4).value
        sheet.cell(row=i,column=15).value = gg_end_lat - sheet_gg.cell(row=i,column=5).value
        
        sheet.cell(row=i,column=17).value = haversine(sheet_gg.cell(row=i,column=4).value,sheet_gg.cell(row=i,column=5).value,gg_end_lon,gg_end_lat) * 1000

sheet_file = "C:\\Users\\phams\\Downloads\\check_google_results.xlsx"
spreadsheet.save(sheet_file)