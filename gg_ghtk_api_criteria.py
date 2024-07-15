import openpyxl
from datetime import date
from openpyxl.styles import PatternFill, Font



today = date.today()

spreadsheet_general = openpyxl.load_workbook(f"C:\\Users\\phams\\Downloads\\linestring_analysis_{today}.xlsx")
sheet_general = spreadsheet_general.active

spreadsheet_start_end = openpyxl.load_workbook(f"C:\\Users\\phams\\Downloads\\linestring_analysis_start_end_{today}.xlsx")
sheet_start_end = spreadsheet_start_end.active

sheet_general.cell(row=1,column=23).value = 'Status car'
sheet_general.cell(row=1,column=24).value = 'Status bike'
sheet_general.cell(row=1,column=25).value = 'Status motor'
sheet_general.cell(row=1,column=26).value = 'Status xteam motor'

sheet_start_end.cell(row=1,column=20).value = 'Status car'
sheet_start_end.cell(row=1,column=21).value = 'Status bike'
sheet_start_end.cell(row=1,column=22).value = 'Status motor'
sheet_start_end.cell(row=1,column=23).value = 'Status xteam motor'

for i in range(2,sheet_general.max_row+1):
    
    for x in range(4):
        if sheet_general.cell(row=i,column=19+x).value != "No error":
            # Return error
            sheet_general.cell(row=i,column=23+x).value = "Error"
            sheet_general.cell(row=i,column=23+x).font = Font(color="a99816") # yellow
            sheet_general.cell(row=i,column=23+x).fill = PatternFill(start_color='ecd41f', end_color='ecd41f', fill_type='solid') # yellow
            
            sheet_start_end.cell(row=i,column=20+x).value = "Error"
            sheet_start_end.cell(row=i,column=20+x).font = Font(color="a99816") # yellow
            sheet_start_end.cell(row=i,column=20+x).fill = PatternFill(start_color='ecd41f', end_color='ecd41f', fill_type='solid') # yellow
        else:
            # Return route
            if sheet_general.cell(row=i,column=11+x).value >= -10 and sheet_general.cell(row=i,column=11+x).value <= 10:
                # Distance difference is no more than 10% of Google route
                if sheet_general.cell(row=i,column=15+x).value >= 80 and sheet_general.cell(row=i,column=15+x).value <= 100:
                    # GHTK route is within Google route's buffer
                    sheet_general.cell(row=i,column=23+x).value = "Acceptable"
                    sheet_general.cell(row=i,column=23+x).font = Font(color="1aa113") # green
                    sheet_general.cell(row=i,column=23+x).fill = PatternFill(start_color='25e91b', end_color='25e91b', fill_type='solid') # green
                else:
                    # GHTK route is not within Google route's buffer
                    sheet_general.cell(row=i,column=23+x).value = "Not in buffer"
                    sheet_general.cell(row=i,column=23+x).font = Font(color="a5190b") # red
                    sheet_general.cell(row=i,column=23+x).fill = PatternFill(start_color='f2230f', end_color='f2230f', fill_type='solid') # red
            else:
                # Distance difference is more than 10% of Google route
                sheet_general.cell(row=i,column=23+x).value = "Too long / short"
                sheet_general.cell(row=i,column=23+x).font = Font(color="a5190b") # red
                sheet_general.cell(row=i,column=23+x).fill = PatternFill(start_color='f2230f', end_color='f2230f', fill_type='solid') # red
    
        if sheet_start_end.cell(row=i,column=6).value > 100 or sheet_start_end.cell(row=i,column=7).value > 100:
            # Distance from Google route starting point to input starting point or from Google route ending point to input ending point is greater than 100m
            sheet_start_end.cell(row=i,column=20+x).value = 'Wrong Google route'
            sheet_start_end.cell(row=i,column=20+x).font = Font(color="a5190b") # red
            sheet_start_end.cell(row=i,column=20+x).fill = PatternFill(start_color='f2230f', end_color='f2230f', fill_type='solid') # red
        else:
            # Distance from Google route starting point to input starting point and from Google route ending point to input ending point is no more than 100m
            if sheet_start_end.cell(row=i,column=8+(2*x)).value > 100 or sheet_start_end.cell(row=i,column=9+(2*x)).value > 100:
                # Distance from corresponding GHTK route starting point to input starting point or from corresponding GHTK route ending point to input ending point is greater than 100m
                sheet_start_end.cell(row=i,column=20+x).value = 'Right Google route, Wrong GHTK route'
                sheet_start_end.cell(row=i,column=20+x).font = Font(color="a5190b") # red
                sheet_start_end.cell(row=i,column=20+x).fill = PatternFill(start_color='f2230f', end_color='f2230f', fill_type='solid') # red
            else:
                # Distance from corresponding GHTK route starting point to input starting point and from corresponding GHTK route ending point to input ending point is no more than 100m
                sheet_start_end.cell(row=i,column=20+x).value = 'Right Google route, Right GHTK route'
                sheet_start_end.cell(row=i,column=20+x).font = Font(color="1aa113") # green
                sheet_start_end.cell(row=i,column=20+x).fill = PatternFill(start_color='25e91b', end_color='25e91b', fill_type='solid') # green

spreadsheet_general.save(f"C:\\Users\\phams\\Downloads\\linestring_analysis_{today}.xlsx")
spreadsheet_start_end.save(f"C:\\Users\\phams\\Downloads\\linestring_analysis_start_end_{today}.xlsx")