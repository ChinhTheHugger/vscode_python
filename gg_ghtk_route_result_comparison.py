import openpyxl
from datetime import date
from openpyxl.styles import PatternFill, Font



today = date.today()

spreadsheet = openpyxl.Workbook()

sheet_car = spreadsheet.create_sheet(title='car')
sheet_bike = spreadsheet.create_sheet(title='bike')
sheet_motor = spreadsheet.create_sheet(title='motor')
sheet_xmotor = spreadsheet.create_sheet(title='xteam motor')

sheets = [sheet_car,sheet_bike,sheet_motor,sheet_xmotor]

spreadsheet_general = openpyxl.load_workbook(f"C:\\Users\\phams\\Downloads\\linestring_analysis_{today}.xlsx")
sheet_general = spreadsheet_general.active

spreadsheet_start_end = openpyxl.load_workbook(f"C:\\Users\\phams\\Downloads\\linestring_analysis_start_end_{today}.xlsx")
sheet_start_end = spreadsheet_start_end.active

for i in range(2,sheet_general.max_row+1):
    for x in range(4):
        sheets[x].cell(row=i,column=1).value = sheet_general.cell(row=i,column=1).value
        sheets[x].cell(row=i,column=2).value = sheet_general.cell(row=i,column=23+x).value
        sheets[x].cell(row=i,column=3).value = 