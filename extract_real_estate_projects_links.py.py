import json
import openpyxl
import os
import time

# spreadsheet = openpyxl.Workbook()
# sheet = spreadsheet.active

json_list = [('phu_tho',1),('ha_giang',1),('cao_bang',1),('bac_kan',1),('tuyen_quang',1),('lao_cai',1),('dien_bien',1)
             ,('lai_chau',1),('son_la',1),('yen_bai',1),('hoa_binh',1),('thai_binh',2),('ha_nam',2),('nam_dinh',1),('ninh_binh',1),('thanh_hoa',2)
             ,('nghe_an',1),('ha_tinh',1),('quang_binh',1),('quang_tri',1),('thua_thien_hue',1),('da_nang',1),('quang_nam',1),('quang_ngai',1),('binh_dinh',1)
             ,('phu_yen',1),('khanh_hoa',1),('ninh_thuan',1),('binh_thuan',1),('kon_tum',1),('gia_lai',1),('dak_lak',1),('dak_nong',1),('lam_dong',1)
             ,('binh_phuoc',2),('tay_ninh',1),('thai_binh_duong',3),('dong_nai',2),('ba_ria_vung_tau',2),('tp_hcm',2),('long_an',2),('tien_giang',1),('ben_tre',1)
             ,('tra_vinh',1),('vinh_long',1),('dong_thap',1),('an_giang',1),('kien_giang',1),('can_tho',1),('hau_giang',1),('soc_trang',1),('bac_lieu',1),('ca_mau',1)]

for pair in json_list:
    name, count = pair[0], pair[1]
    
    parts = name.split('_')
    capitalized_parts = [part.capitalize() for part in parts]
    
    capitalized_name = '_'.join(capitalized_parts)
    
    new_folder_path = f'C:\\Users\\phams\\Downloads\\{capitalized_name}'
    os.makedirs(new_folder_path, exist_ok=True)
    
    new_folder_path_source = f'C:\\Users\\phams\\Downloads\\{capitalized_name}\\page sources'
    os.makedirs(new_folder_path_source, exist_ok=True)
    
    new_folder_path_info = f'C:\\Users\\phams\\Downloads\\{capitalized_name}\\thong tin du an'
    os.makedirs(new_folder_path_info, exist_ok=True)
    
    # spreadsheet = openpyxl.load_workbook("C:\\Users\\phams\\Downloads\\Hung_Yen\\hung_yen_projects_links.xlsx")
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active
    
    links = []

    for i in range(count):
        file_path = f"C:\\Users\\phams\\Downloads\\{name}_page_{i+1}.json"

        # Attempt to open the file with 'utf-8' encoding
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
        except UnicodeDecodeError:
            # If 'utf-8' encoding fails, try 'iso-8859-1'
            with open(file_path, 'r', encoding='iso-8859-1') as file:
                data = json.load(file)

        if len(data['children']) != 0:
            for i in range(len(data['children'])):
                if len(data['children'][i]['children']) != 0:
                    if data['children'][i]['children'][0]['value'] not in links and data['children'][i]['children'][0]['value'] != '':
                        links.append(data['children'][i]['children'][0]['value'])

    # print(len(links))

    sheet.cell(row=1,column=1).value = 'Page source'
    sheet.cell(row=1,column=2).value = 'Links'
    sheet.cell(row=1,column=3).value = 'Project name'

    for x in range(len(links)):
        sheet.cell(row=x+2,column=1).value = x+1
        sheet.cell(row=x+2,column=2).value = links[x]
        
    sheet_file = f"C:\\Users\\phams\\Downloads\\{capitalized_name}\\{name}_projects_links.xlsx"
    spreadsheet.save(sheet_file)
    
    print(f'Finished extracting: {name}')

