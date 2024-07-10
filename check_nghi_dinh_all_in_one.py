import asyncio
from playwright.async_api import async_playwright
import json
from datetime import date
import openpyxl
import os
import re
from openpyxl.styles import Alignment

path = 'C:\\Users\\phams\\Downloads\\nghi dinh\\nghi_dinh.xlsx'

if not os.path.exists(path):
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active
    
    sheet.cell(row=1,column=1).value = 'Num.'
    sheet.cell(row=1,column=2).value = 'Date'
    sheet.cell(row=1,column=3).value = 'JSON'
    sheet.cell(row=1,column=4).value = 'Sheet'
    sheet.cell(row=1,column=5).value = 'File name'
    
    spreadsheet.save(path)

spreadsheet = openpyxl.load_workbook(path)
sheet = spreadsheet.active

async def get_accessibility_tree(url):
    async with async_playwright() as p:
        browser = await p.firefox.launch_persistent_context(user_data_dir="C:\\Users\\phams\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\bd5r7ma0.default-release - Copy", headless=True)
        page = await browser.new_page()
        await page.goto(url)
        
        # Get the accessibility tree
        accessibility_snapshot = await page.accessibility.snapshot()
        
        today = date.today()
        
        with open(f'C:\\Users\\phams\\Downloads\\nghi dinh\\json\\nghi_dinh_list_{today}.json', 'w') as f:
            json.dump(accessibility_snapshot, f, indent=2)
        
        rows = sheet.max_row
        
        if sheet.cell(row=rows,column=3).value != f'json\\nghi_dinh_list_{today}.json':
            sheet.cell(row=rows+1,column=1).value = rows
            sheet.cell(row=rows+1,column=2).value = today
            sheet.cell(row=rows+1,column=3).value = f'json\\nghi_dinh_list_{today}.json'
            sheet.cell(row=rows+1,column=4).value = f'sheet\\nghi_dinh_list_{today}.xlsx'
        
        print(f'Decree lists, date: {today}')
        
        await browser.close()
        
        # # Get the page source
        # page_source = await page.content()
        
        # # Save the page source to an HTML file
        # with open('C:\\Users\\phams\\Downloads\\page_source.html', 'w', encoding='utf-8') as f:
        #     f.write(page_source)
        
        # await browser.close()

# URL to open
url = 'https://danhmuchanhchinh.gso.gov.vn/NghiDinh.aspx'

# Run the function
asyncio.get_event_loop().run_until_complete(get_accessibility_tree(url))

spreadsheet.save(path)

spreadsheet = openpyxl.load_workbook(path)
sheet = spreadsheet.active

rows = sheet.max_row

json_path_last_version = f'C:\\Users\\phams\\Downloads\\nghi dinh\\json\\{sheet.cell(row=rows-1,column=5).value}.json'
sheet_path_last_version = f'C:\\Users\\phams\\Downloads\\nghi dinh\\sheet\\{sheet.cell(row=rows-1,column=5).value}.xlsx'

json_path = f'C:\\Users\\phams\\Downloads\\nghi dinh\\json\\{sheet.cell(row=rows,column=5).value}.json'
sheet_path = f'C:\\Users\\phams\\Downloads\\nghi dinh\\sheet\\{sheet.cell(row=rows,column=5).value}.xlsx'

# Attempt to open the file with 'utf-8' encoding
try:
    with open(json_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
except UnicodeDecodeError:
    # If 'utf-8' encoding fails, try 'iso-8859-1'
    with open(json_path, 'r', encoding='iso-8859-1') as file:
        data = json.load(file)

spreadsheet_list = openpyxl.Workbook()
sheet_list = spreadsheet_list.active

sheet_list.cell(row=1,column=1).value = 'Num.'
sheet_list.cell(row=1,column=2).value = 'Nghi dinh so'
sheet_list.cell(row=1,column=3).value = 'Ngay ban hanh'
sheet_list.cell(row=1,column=4).value = 'Ngay hieu luc'
sheet_list.cell(row=1,column=5).value = 'Noi dung'
sheet_list.cell(row=1,column=6).value = 'Status'

# Define the column widths
column_widths = {'A': 8, 'B': 25, 'C': 20, 'D': 20, 'E': 100}

# Set the column widths
for col, width in column_widths.items():
    sheet_list.column_dimensions[col].width = width
    
spreadsheet_last_version = openpyxl.load_workbook(sheet_path_last_version)
sheet_last_version = spreadsheet_last_version.active

decree_codes = [cell.value for cell in sheet_last_version['B']]

if len(data['children']) != 0:
    decrees = data['children'][-2]['children']
    
    # print(len(decrees))
    
    # decree_count = data['children'][-1]['name']
    
    # # Use regex to find the first number in the string
    # match = re.search(r'\d+', decree_count)

    # # Extract the number (if found) and convert it to an integer
    # if match:
    #     number = int(match.group())
    #     print(number)
    # else:
    #     print("No number found in the string.")
    
    # total_children = number * 4
    
    # for i in range(total_children):
    #     count = 4 * i
    #     for x in range(4):
    #         sheet_list.cell(row=i+2,column=x+1).value = decrees[x+count]['name']
    
    total = int(len(decrees) / 4)
    
    for i in range(total):
        sheet_list.cell(row=i+2,column=1).value = i+1
        for x in range(4):
            sheet_list.cell(row=i+2,column=x+2).value = decrees[x + (4 * i)]['name']
            if x == 0:
                decree_code = sheet_list.cell(row=i+2,column=x+2).value
                if decree_code not in decree_codes:
                    sheet_list.cell(row=i+2,column=6).value = 'NEW'
            if x+2 == 5:
                sheet_list.cell(row=i+2,column=x+2).alignment = Alignment(wrapText=True)

spreadsheet_list.save(sheet_path)

print(f'Saved new decree list on {date.today()}')