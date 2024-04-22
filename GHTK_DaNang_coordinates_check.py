import openpyxl
import pandas
from collections import Counter
from difflib import SequenceMatcher
from collections import OrderedDict
import time
import numpy
import igraph
import time
from datetime import datetime
import haversine
from haversine import Unit
from math import radians, sin, cos, acos
from geopy.distance import geodesic
from geopy.distance import great_circle
from openpyxl.chart import BarChart, series, Reference



path_true_coordinates = "C:\\Users\\phams\\Downloads\\Đà_Nẵng_Toạ_độ_đúng.XLSX"
path_ip1_ip2_coordinates = "C:\\Users\\phams\\Downloads\\Dataraw_đánh_giá_chất_lượng_phân_tích_địa_chỉ_tại_Đà_Nẵng.XLSX"
path_result = "C:\\Users\\phams\\Downloads\\DaNang_ip_check.XLSX"

wb_true = openpyxl.load_workbook(path_true_coordinates)
sheet_true = wb_true.active

wb_ip = openpyxl.load_workbook(path_ip1_ip2_coordinates)
sheet_ip = wb_ip.active

wb_result = openpyxl.load_workbook(path_result)
sheet_result = wb_result.active

start_calculation = time.time()
start = datetime.now()
# print("Start time = " + time.strftime("%H:%M:%S", time.gmtime(start)))
print("Start time = ", start)
print("Delay: 2 seconds")
time.sleep(2)



# # populate path_result with data
# for i in range(2,7702):
#     sheet_result.cell(row=i,column=1).value = i - 1
#     sheet_result.cell(row=i,column=2).value = sheet_true.cell(row=i,column=2).value
#     sheet_result.cell(row=i,column=3).value = sheet_true.cell(row=i,column=3).value
#     sheet_result.cell(row=i,column=4).value = sheet_true.cell(row=i,column=4).value
#     sheet_result.cell(row=i,column=5).value = sheet_ip.cell(row=i,column=7).value
#     sheet_result.cell(row=i,column=6).value = sheet_ip.cell(row=i,column=9).value
#     sheet_result.cell(row=i,column=7).value = sheet_ip.cell(row=i,column=11).value
#     sheet_result.cell(row=i,column=8).value = sheet_ip.cell(row=i,column=2).value
#     sheet_result.cell(row=i,column=9).value = sheet_ip.cell(row=i,column=3).value
#     sheet_result.cell(row=i,column=10).value = sheet_ip.cell(row=i,column=4).value
#     sheet_result.cell(row=i,column=11).value = sheet_true.cell(row=i,column=6).value
#     sheet_result.cell(row=i,column=12).value = sheet_true.cell(row=i,column=7).value
#     sheet_result.cell(row=i,column=13).value = sheet_ip.cell(row=i,column=14).value
#     sheet_result.cell(row=i,column=14).value = sheet_ip.cell(row=i,column=15).value
# wb_result.save(path_result)



# # adjust row and column size
# sheet_result.column_dimensions['B'].width = 15
# sheet_result.column_dimensions['C'].width = 30
# sheet_result.column_dimensions['D'].width = 30
# sheet_result.column_dimensions['E'].width = 15
# sheet_result.column_dimensions['F'].width = 30
# sheet_result.column_dimensions['G'].width = 30
# sheet_result.column_dimensions['H'].width = 15
# sheet_result.column_dimensions['I'].width = 30
# sheet_result.column_dimensions['J'].width = 30
# sheet_result.row_dimensions[1].height = 30
# wb_result.save(path_result)



# # dcc comparison
# for i in range(2,7702):
#     if sheet_result.cell(row=i,column=2).value != sheet_result.cell(row=i,column=5).value:
#         sheet_result.cell(row=i,column=15).value = "X"
#     else:
#         sheet_result.cell(row=i,column=15).value = ""
        
#     if sheet_result.cell(row=i,column=3).value != sheet_result.cell(row=i,column=6).value:
#         sheet_result.cell(row=i,column=17).value = "X"
#     else:
#         sheet_result.cell(row=i,column=17).value = ""
        
#     if sheet_result.cell(row=i,column=4).value != sheet_result.cell(row=i,column=7).value:
#         sheet_result.cell(row=i,column=19).value = "X"
#     else:
#         sheet_result.cell(row=i,column=19).value = ""
        
#     if sheet_result.cell(row=i,column=2).value != sheet_result.cell(row=i,column=8).value:
#         sheet_result.cell(row=i,column=16).value = "X"
#     else:
#         sheet_result.cell(row=i,column=16).value = ""
        
#     if sheet_result.cell(row=i,column=3).value != sheet_result.cell(row=i,column=9).value:
#         sheet_result.cell(row=i,column=18).value = "X"
#     else:
#         sheet_result.cell(row=i,column=18).value = ""
        
#     if sheet_result.cell(row=i,column=4).value != sheet_result.cell(row=i,column=10).value:
#         sheet_result.cell(row=i,column=20).value = "X"
#     else:
#         sheet_result.cell(row=i,column=20).value = ""
# wb_result.save(path_result) 



# # coordinate distance
# for i in range(2,7702):
#     old_coordinates = (float(sheet_result.cell(row=i,column=13).value),float(sheet_result.cell(row=i,column=14).value))
#     true_coordinates = (float(sheet_result.cell(row=i,column=11).value),float(sheet_result.cell(row=i,column=12).value))
    
#     # haversine
#     sheet_result.cell(row=i,column=21).value = haversine.haversine(old_coordinates,true_coordinates,unit=Unit.METERS)
    
#     # math module
#     old_lat = float(sheet_result.cell(row=i,column=13).value)
#     old_long = float(sheet_result.cell(row=i,column=14).value)
#     true_lat = float(sheet_result.cell(row=i,column=11).value)
#     true_long = float(sheet_result.cell(row=i,column=12).value)
#     # earth radius = 6371.01 km
#     sheet_result.cell(row=i,column=22).value = 6371010 * acos(sin(old_lat)*sin(true_lat) + cos(old_lat)*cos(true_lat)*cos(old_long - true_long))
    
#     # geodesic
#     sheet_result.cell(row=i,column=23).value = geodesic(old_coordinates,true_coordinates).m
    
#     # great circle formula
#     sheet_result.cell(row=i,column=24).value = great_circle(old_coordinates,true_coordinates).m
# wb_result.save(path_result)

# # NOTES:
# # - math module answer is not correct
# # - haversine, geodesic, great circle return similar values
# # - haversine and great circle results are more close together


# get min, max values
dataHaversine = set(sheet_result.cell(row=i,column=21).value for i in range(2,7702))
dataGeodesic = set(sheet_result.cell(row=i,column=23).value for i in range(2,7702))
dataGreatCircle = set(sheet_result.cell(row=i,column=24).value for i in range(2,7702))
print("Haversine: min = " + str(min(dataHaversine)) + "m, max = " + str(max(dataHaversine)) + "m\n")
print("Geodesic: min = "+ str(min(dataGeodesic)) + "m, max = "+ str(max(dataGeodesic)) + "m\n")
print("Great Circle: min = "+ str(min(dataGreatCircle)) + "m, max = "+ str(max(dataGreatCircle)) + "m\n")



# # export data to bar chart
# chartHaversine = BarChart()
# chartHaversine.height = 30
# chartHaversine.width = 50
# chartHaversine.type = "col"
# chartHaversine.style = 10
# chartHaversine.title = "Haversine"
# chartHaversine.y_axis.title = "Count"
# chartHaversine.x_axis.title = "Distances"

# values = Reference(sheet_result,min_col=21,min_row=2,max_col=21,max_row=7701)
# chartHaversine.add_data(values)
# sheet_result.add_chart(chartHaversine,"Z2")
# wb_result.save(path_result)



# value range percentage

# haversine
sub50_count = 0
sub100_count = 0
sub500_count = 0
sub1000_count = 0
above1000_count = 0
for i in range(2,7702):
    if sheet_result.cell(row=i,column=21).value < 50:
        sub50_count +=1
    elif sheet_result.cell(row=i,column=21).value >= 50 and sheet_result.cell(row=i,column=21).value < 100:
        sub100_count += 1
    elif sheet_result.cell(row=i,column=21).value >=100 and sheet_result.cell(row=i,column=21).value<500:
        sub500_count += 1
    elif sheet_result.cell(row=i,column=21).value >= 500 and sheet_result.cell(row=i,column=21).value < 1000:
        sub1000_count += 1
    elif sheet_result.cell(row=i,column=21).value >= 1000:
        above1000_count += 1
print("Haversine:\n- 0 to 50m: "+str(100*(float(sub50_count)/float(7700)))+"%\n- 50 to 100m: "+str(100*(float(sub100_count)/float(7700)))+"%\n- 100 to 500m: "+str(100*(float(sub500_count)/float(7700)))+"%\n- 500 to 1000m: "+str(100*(float(sub1000_count)/float(7700)))+"%\n- 1000m and higher: "+str(100*(float(above1000_count)/float(7700)))+"%\n")

# geodesic
sub50_count = 0
sub100_count = 0
sub500_count = 0
sub1000_count = 0
above1000_count = 0
for i in range(2,7702):
    if sheet_result.cell(row=i,column=23).value < 50:
        sub50_count +=1
    elif sheet_result.cell(row=i,column=23).value >= 50 and sheet_result.cell(row=i,column=23).value < 100:
        sub100_count += 1
    elif sheet_result.cell(row=i,column=23).value >=100 and sheet_result.cell(row=i,column=23).value<500:
        sub500_count += 1
    elif sheet_result.cell(row=i,column=23).value >= 500 and sheet_result.cell(row=i,column=23).value < 1000:
        sub1000_count += 1
    elif sheet_result.cell(row=i,column=23).value >= 1000:
        above1000_count += 1
print("Geodesic:\n- 0 to 50m: "+str(100*(float(sub50_count)/float(7700)))+"%\n- 50 to 100m: "+str(100*(float(sub100_count)/float(7700)))+"%\n- 100 to 500m: "+str(100*(float(sub500_count)/float(7700)))+"%\n- 500 to 1000m: "+str(100*(float(sub1000_count)/float(7700)))+"%\n- 1000m and higher: "+str(100*(float(above1000_count)/float(7700)))+"%\n")

# great circle
sub50_count = 0
sub100_count = 0
sub500_count = 0
sub1000_count = 0
above1000_count = 0
for i in range(2,7702):
    if sheet_result.cell(row=i,column=24).value < 50:
        sub50_count +=1
    elif sheet_result.cell(row=i,column=24).value >= 50 and sheet_result.cell(row=i,column=24).value < 100:
        sub100_count += 1
    elif sheet_result.cell(row=i,column=24).value >=100 and sheet_result.cell(row=i,column=24).value<500:
        sub500_count += 1
    elif sheet_result.cell(row=i,column=24).value >= 500 and sheet_result.cell(row=i,column=24).value < 1000:
        sub1000_count += 1
    elif sheet_result.cell(row=i,column=24).value >= 1000:
        above1000_count += 1
print("Great circle:\n- 0 to 50m: "+str(100*(float(sub50_count)/float(7700)))+"%\n- 50 to 100m: "+str(100*(float(sub100_count)/float(7700)))+"%\n- 100 to 500m: "+str(100*(float(sub500_count)/float(7700)))+"%\n- 500 to 1000m: "+str(100*(float(sub1000_count)/float(7700)))+"%\n- 1000m and higher: "+str(100*(float(above1000_count)/float(7700)))+"%\n")

# NOTES:
# - all 3 methods have 54.5% of distance values under 50m



end_calculation = time.time()
end = datetime.now()

# print("End time = " + time.strftime("%H:%M:%S", time.gmtime(end)))
print("End time = ", end)
print("Total run time = ", time.strftime("%H:%M:%S", time.gmtime(end_calculation-start_calculation)))