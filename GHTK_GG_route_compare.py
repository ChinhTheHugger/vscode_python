import openpyxl
import polyline
from shapely import LineString
import numpy as np
import math
import requests



spreadsheet_gg = openpyxl.load_workbook("C:\\Users\\phams\\Downloads\\failed_case.xlsx")
sheet_gg = spreadsheet_gg.active

spreadsheet_rs = openpyxl.load_workbook("C:\\Users\\phams\\Downloads\\check_google_results(1).xlsx")
sheet_rs = spreadsheet_rs.active

# result sheet
spreadsheet = openpyxl.Workbook()
sheet = spreadsheet.active



# Haversine distance function - return in kilometer
def haversine(lon1, lat1, lon2, lat2):
    # Radius of the Earth in km
    R = 6371.0

    # Convert decimal degrees to radians
    lat1 = math.radians(lat1)
    lon1 = math.radians(lon1)
    lat2 = math.radians(lat2)
    lon2 = math.radians(lon2)

    # Differences in coordinates
    dlat = lat2 - lat1
    dlon = lon2 - lon1

    # Haversine formula
    a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    # Distance in kilometers
    distance = R * c

    return distance

def ghtk_gh_api_local(start,end,type):
    headers = {
        'apikey': 'aghzzzzzzzzzzzzui',
        'Content-Type': 'application/json',
        'Authorization': 'Bearer eyJhbGciOiJFUzI1NiIsImtpZCI6IjAxRjRTMFBTRjdQR001VDZDWVExTTVNRVg3XzE2MjAwNDIyNzgiLCJ0eXAiOiJKV1QifQ.eyJhdWQiOiJnbWFwIiwiZXhwIjoxNjcwMzg4OTg0LCJqdGkiOiIwMUdLSldOS1NTN0FDSldTTVExMURHNzhZRSIsImlhdCI6MTY3MDMwMjU3OSwiaXNzIjoiaHR0cHM6Ly9hdXRoLmdodGtsYWIuY29tIiwic3ViIjoiMDFGNFMwUFNGN1BHTTVUNkNZUTFNNU1FWDciLCJzY3AiOlsiZ21hcDphZGRyZXNzLXZlcmlmaWNhdGlvbi1hcGkuYWNjZXNzLWRpcmVjdCIsImdtYXA6YWRkcmVzc2VzLW1hbmFnZW1lbnQuYWNjZXNzLWRpcmVjdCIsImdtYXA6Y2FwYWNpdHkuYWNjZXNzLWRpcmVjdCIsImdtYXA6Z29vZ2xlLW1hcC1odHRwLXNlcnZpY2UuYWNjZXNzLWRpcmVjdCIsImdtYXA6aGFtbGV0LW5vZGUtZGV0YWlsLmFjY2Vzcy1kaXJlY3QiLCJnbWFwOmludGVybmFsLWFwaS5hY2Nlc3MtZGlyZWN0IiwiZ21hcDpwcmVnZW8uYWNjZXNzLWRpcmVjdCIsImdtYXA6cm91dGUuYWNjZXNzLWRpcmVjdCIsImdtYXA6c3VnZ2VzdC1hZGRyZXNzLmFjY2Vzcy1kaXJlY3QiXSwiY2xpZW50X2lkIjoiMDFGNFMwUFNGN1BHTTVUNkNZUTFNNU1FWDciLCJ0eXBlIjoiZGlyZWN0In0.b89nX-mK6pjtfAMtqcCXFghvPC1qbOHjPDaDDrz9ljuTpdAhkvxkVlnST2tSyxNJ4dluwjHIBPsmFzWS5erJUA',
    }

    json_data = {
        'gh_requests': [
            {
                'points': [
                    [
                        start[1],start[0]
                    ],
                    [
                        end[1],end[0]
                    ],
                ],
                'profile': type,
                'request_id': '982884285',
                'calc_points': True,
                'points_encoded': True,
                'instructions': True,
            },
        ],
    }

    response = requests.post('http://10.110.69.186:8989/route', headers=headers, json=json_data)

    # Note: json_data will not be serialized by requests
    # exactly as it was in the original request.
    #data = '{\n        "gh_requests": [\n            {\n                "points": [\n                    [\n                        105.842638,\n                        21.200526\n                    ],\n                    [\n                        104.1166,\n                        21.185161\n                    ]\n                ],\n                "profile": "car",\n                "request_id": "982884285",\n                "calc_points": true,\n                "points_encoded": true,\n                "instructions": true\n            }\n        ]\n    }'
    #response = requests.post('http://10.110.69.186:8989/route', headers=headers, data=data)

    data = response.json()
    
    if data.get('gh_responses') is None:
        data = {
            'status': False,
            'route': '',
            'message': "Error getting response"
        }
        
        return data
    else:
        gh_response = data['gh_responses']
        if gh_response[0].get('paths', None) is None or gh_response[0]['paths'] == []:
            data = {
                'status': False,
                'route': '',
                'message': gh_response[0]['errors'][0]['message']
            }
            
            return data
        else:
            paths_data = gh_response[0]['paths']
            encoded_points = paths_data[0]['points']
            decoded_points = polyline.decode(encoded_points)
            
            data = {
                'status': True,
                'route': [(lon, lat) for lat, lon in decoded_points],
                'message': "No error"
            }
            
            return data



sheet.cell(row=1,column=1).value = 'Case'

sheet.cell(row=1,column=2).value = 'Start long - original'
sheet.cell(row=1,column=3).value = 'Start lat - original'

sheet.cell(row=1,column=4).value = 'End long - original'
sheet.cell(row=1,column=5).value = 'End lat - original'

sheet.cell(row=1,column=6).value = 'Start long - from GG'
sheet.cell(row=1,column=7).value = 'Start lat - from GG'

sheet.cell(row=1,column=8).value = 'End long - from GG'
sheet.cell(row=1,column=9).value = 'End lat - from GG'

sheet.cell(row=1,column=10).value = 'Start long - from GHTK'
sheet.cell(row=1,column=11).value = 'Start lat - from GHTK'

sheet.cell(row=1,column=12).value = 'End long - from GHTK'
sheet.cell(row=1,column=13).value = 'End lat - from GHTK'

sheet.cell(row=1,column=14).value = 'Start points haversine distance - GG (m)'
sheet.cell(row=1,column=15).value = 'End points haversine distance - GG (m)'

sheet.cell(row=1,column=16).value = 'Start points haversine distance - GHTK (m)'
sheet.cell(row=1,column=17).value = 'End points haversine distance - GHTK (m)'

sheet.cell(row=1,column=18).value = 'GHTK error'

for i in range(2, sheet_rs.max_row+1):
    sheet.cell(row=i,column=1).value = sheet_rs.cell(row=i,column=1).value
    
    sheet.cell(row=i,column=2).value = sheet_rs.cell(row=i,column=2).value
    sheet.cell(row=i,column=3).value = sheet_rs.cell(row=i,column=3).value
    
    sheet.cell(row=i,column=4).value = sheet_rs.cell(row=i,column=4).value
    sheet.cell(row=i,column=5).value = sheet_rs.cell(row=i,column=5).value
    
    start = (sheet_rs.cell(row=i,column=3).value,sheet_rs.cell(row=i,column=2).value) # lat, long
    end = (sheet_rs.cell(row=i,column=5).value,sheet_rs.cell(row=i,column=4).value) # lat, long
    print(start)
    sheet.cell(row=i,column=6).value = sheet_rs.cell(row=i,column=6).value
    sheet.cell(row=i,column=7).value = sheet_rs.cell(row=i,column=7).value
            
    sheet.cell(row=i,column=8).value = sheet_rs.cell(row=i,column=8).value
    sheet.cell(row=i,column=9).value = sheet_rs.cell(row=i,column=9).value
    
    sheet.cell(row=i,column=14).value = sheet_rs.cell(row=i,column=10).value
    sheet.cell(row=i,column=15).value = sheet_rs.cell(row=i,column=11).value
            
    gh_response_motor = ghtk_gh_api_local(start,end,'motorcycle')
            
    if gh_response_motor['status'] is False:
        gh_line_motor = LineString([])
        error_message_motor = gh_response_motor['message']
    else:
        gh_line_motor = LineString(gh_response_motor['route'])
        error_message_motor = gh_response_motor['message']
    
    print(error_message_motor)
            
    if not gh_line_motor.is_empty:
        gh_start_lon, gh_start_lat = gh_line_motor.coords[0]
        gh_end_lon, gh_end_lat = gh_line_motor.coords[-1]
                
        sheet.cell(row=i,column=10).value = gh_start_lon
        sheet.cell(row=i,column=11).value = gh_start_lat
                
        sheet.cell(row=i,column=12).value = gh_end_lon
        sheet.cell(row=i,column=13).value = gh_end_lat
                
        sheet.cell(row=i,column=16).value = haversine(sheet_rs.cell(row=i,column=2).value,sheet_rs.cell(row=i,column=3).value,gh_start_lon,gh_start_lat) * 1000
        sheet.cell(row=i,column=17).value = haversine(sheet_rs.cell(row=i,column=4).value,sheet_rs.cell(row=i,column=5).value,gh_end_lon,gh_end_lat) * 1000
                
        sheet.cell(row=i,column=18).value = error_message_motor
    else:
        sheet.cell(row=i,column=10).value = 0
        sheet.cell(row=i,column=11).value = 0
                
        sheet.cell(row=i,column=12).value = 0
        sheet.cell(row=i,column=13).value = 0
                
        sheet.cell(row=i,column=16).value = 0
        sheet.cell(row=i,column=17).value = 0
                
        sheet.cell(row=i,column=18).value = error_message_motor

sheet_file = "C:\\Users\\phams\\Downloads\\check_google_results_v2.xlsx"
spreadsheet.save(sheet_file)