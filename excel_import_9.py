import openpyxl
from collections import Counter
import time
import numpy

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.17 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing - H+P IDs.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()

# Remove house dup

# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+1):
#         if sheet_obj.cell(row=i,column=4).value != sheet_obj.cell(row=x,column=4).value:
#             break
#         else:
#             sheet_obj.cell(row=i,column=2).value = None
#     x=i

# Remove phone dup

# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+1):
#         if sheet_obj.cell(row=i,column=5).value != sheet_obj.cell(row=x,column=5).value:
#             break
#         else:
#             sheet_obj.cell(row=i,column=3).value = None
#     x=i

# Remove Id dup

for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,sheet_obj.max_row+1):
        if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
            break
        else:
            sheet_obj.cell(row=i,column=1).value = None
    x=i



wb_obj.save(path)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))