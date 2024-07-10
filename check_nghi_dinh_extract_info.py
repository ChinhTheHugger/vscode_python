import asyncio
from playwright.async_api import async_playwright
import json
from datetime import date
import openpyxl
import os
import re
from openpyxl.styles import Alignment

path = 'C:\\Users\\phams\\Downloads\\nghi dinh\\nghi_dinh.xlsx'

spreadsheet = openpyxl.load_workbook(path)
sheet = spreadsheet.active

rows = sheet.max_row

json_path = f'C:\\Users\\phams\\Downloads\\nghi dinh\\json\\{sheet.cell(row=rows,column=5).value}.json'
sheet_path = f'C:\\Users\\phams\\Downloads\\nghi dinh\\sheet\\{sheet.cell(row=rows,column=5).value}.xlsx'

# Attempt to open the file with 'utf-8' encoding
try:
    with open(json_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
except UnicodeDecodeError:
    # If 'utf-8' encoding fails, try 'iso-8859-1'
    with open(json_path, 'r', encoding='iso-8859-1') as file:
        data = json.load(file)

spreadsheet_list = openpyxl.Workbook()
sheet_list = spreadsheet_list.active

sheet_list.cell(row=1,column=1).value = 'Num.'
sheet_list.cell(row=1,column=2).value = 'Nghi dinh so'
sheet_list.cell(row=1,column=3).value = 'Ngay ban hanh'
sheet_list.cell(row=1,column=4).value = 'Ngay hieu luc'
sheet_list.cell(row=1,column=5).value = 'Noi dung'
sheet_list.cell(row=1,column=6).value = 'Status'

# Define the column widths
column_widths = {'A': 8, 'B': 25, 'C': 20, 'D': 20, 'E': 100}

# Set the column widths
for col, width in column_widths.items():
    sheet_list.column_dimensions[col].width = width

if len(data['children']) != 0:
    decrees = data['children'][-2]['children']
    
    # print(len(decrees))
    
    # decree_count = data['children'][-1]['name']
    
    # # Use regex to find the first number in the string
    # match = re.search(r'\d+', decree_count)

    # # Extract the number (if found) and convert it to an integer
    # if match:
    #     number = int(match.group())
    #     print(number)
    # else:
    #     print("No number found in the string.")
    
    # total_children = number * 4
    
    # for i in range(total_children):
    #     count = 4 * i
    #     for x in range(4):
    #         sheet_list.cell(row=i+2,column=x+1).value = decrees[x+count]['name']
    
    total = int(len(decrees) / 4)
    
    for i in range(total):
        sheet_list.cell(row=i+2,column=1).value = i+1
        for x in range(4):
            sheet_list.cell(row=i+2,column=x+2).value = decrees[x + (4 * i)]['name']
            if x+2 == 5:
                sheet_list.cell(row=i+2,column=x+2).alignment = Alignment(wrapText=True)

spreadsheet_list.save(sheet_path)