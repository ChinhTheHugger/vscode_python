import asyncio
from playwright.async_api import async_playwright
import openpyxl
from bs4 import BeautifulSoup
import re
from docx import Document
import os
import time
import json
import pickle

provinces = [
    'an_giang','ba_ria_vung_tau','bac_kan','bac_lieu','bac_ninh','ben_tre','binh_dinh','binh_phuoc','binh_thuan','ca_mau','can_tho','cao_bang','da_nang','dak_lak','dak_nong','dien_bien','dong_nai','dong_thap',
    'gia_lai','ha_giang','ha_nam','ha_tinh','hai_duong','hau_giang','hoa_binh','hung_yen','khanh_hoa','kien_giang','kon_tum','lai_chau','lam_dong','lang_son','lao_cai','long_an','nam_dinh','nghe_an','ninh_binh',
    'ninh_thuan','phu_tho','phu_yen','quang_binh','quang_nam','quang_ngai','quang_tri','soc_trang','son_la','tay_ninh','thai_binh','thai_binh_duong','thai_nguyen','thanh_hoa','thua_thien_hue','tien_giang','tp_hcm',
    'tra_vinh','tuyen_quang','vinh_long','vinh_phuc','yen_bai'
]

for idx, obj in enumerate(provinces):
    name = obj
    
    parts = name.split('_')
    capitalized_parts = [part.capitalize() for part in parts]
    
    capitalized_name = '_'.join(capitalized_parts)
    
    spreadsheet = openpyxl.load_workbook(f'C:\\Users\\phams\\Downloads\\du an\\du lieu goc\\{capitalized_name}\\{obj}_projects_links.xlsx')
    sheet = spreadsheet.active
    
    spreadsheet_check = openpyxl.Workbook()
    sheet_check = spreadsheet_check.active
    
    sheet_check.cell(row=1,column=1).value = 'Page source'
    sheet_check.cell(row=1,column=2).value = 'Project name'
    sheet_check.cell(row=1,column=3).value = '<tbody> in <article>'
    sheet_check.cell(row=1,column=4).value = '<tobdy>'
    sheet_check.cell(row=1,column=5).value = '<article>'
    
    for i in range(2,sheet.max_row+1):
        sheet_check.cell(row=i,column=1).value = sheet.cell(row=i,column=1).value
        sheet_check.cell(row=i,column=2).value = sheet.cell(row=i,column=3).value
        
        with open(f"C:\\Users\\phams\\Downloads\\du an\\du lieu goc\\{capitalized_name}\\page sources\\page_source_{sheet.cell(row=i,column=1).value}.html", 'r', encoding='utf-8') as file:
            html_content = file.read()

        soup = BeautifulSoup(html_content, 'html.parser')

        # detail_article = soup.find('article', class_='detail')
        # if detail_article:
        #     h1_tag = detail_article.find('h1')
        #     if h1_tag:
        #         print(h1_tag.get_text(separator='\n').strip())
        #         print('---------------------------------------------------------------------------')

        article_str = ''
        for tag in soup.find_all('article'):
            string = str(tag.get_text(separator='\n').strip())
            arr = string.split('\n')
            filtered_list = [item.strip() for item in arr if item != '' and item != '\xa0' and item != ' ' and item.strip() != ':'and item.strip() != '']
            # print(filtered_list)
            # print(len(filtered_list))
            article_str = '&'.join(filtered_list)
            sheet_check.cell(row=i,column=5).value = article_str
            # print('---------------------------------------------------------------------------')

        tbody_str  =''
        for tag in soup.find_all('tbody'):
            string = str(tag.get_text(separator='\n').strip())
            arr = string.split('\n')
            filtered_list = [item.strip() for item in arr if item != '' and item != '\xa0' and item != ' ' and item.strip() != ':']
            # print(filtered_list)
            # print(len(filtered_list))
            tbody_str = '&'.join(filtered_list)
            sheet_check.cell(row=i,column=4).value = tbody_str
            # print('---------------------------------------------------------------------------')

        if tbody_str in article_str:
            sheet_check.cell(row=i,column=3).value = 'YES'
        else:
            sheet_check.cell(row=i,column=3).value = 'NO'
    
    spreadsheet_check.save(f'C:\\Users\\phams\\Downloads\\du an\\du lieu goc\\{capitalized_name}\\{obj}_check_data.xlsx')
    
    print(f'Finished checking data for {obj}')
    
    