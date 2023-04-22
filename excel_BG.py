import openpyxl
from collections import Counter
import time
import numpy

path_1 = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\area measurement.XLSX"

wb_1 = openpyxl.load_workbook(path_1)
sheet_1 = wb_1.active

start = time.time()

# Add status

# wb_2 = openpyxl.load_workbook(path_2)
# sheet_2 = wb_2.active

# for x in range(4,sheet_2.max_row+1):
#     if sheet_2.cell(row=x,column=16).value == "x":
#         for i in range(2,sheet_1.max_row+1):
#             if sheet_1.cell(row=i,column=2).value == sheet_2.cell(row=x,column=2).value:
#                 sheet_1.cell(row=i,column=4).value = "Đã bàn giao"

# Add area

# wb_2 = openpyxl.load_workbook(path_3)
# sheet_2 = wb_2.active

# for x in range(2,sheet_1.max_row+1):
#     for i in range(2,sheet_2.max_row+1):
#         if sheet_2.cell(row=i,column=1).value == sheet_1.cell(row=x,column=2).value:
#             sheet_1.cell(row=x,column=4).value = sheet_2.cell(row=i,column=2).value

for x in range(2,sheet_1.max_row+1):
    for i in range(x+1,sheet_1.max_row+2):
        if sheet_1.cell(row=i,column=1).value == None:
            break
        else:
            if sheet_1.cell(row=i,column=2).value == sheet_1.cell(row=x,column=2).value:
                sheet_1.cell(row=i,column=1).value == None
    x=i

wb_1.save(path_1)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))