# DOES NOT WORK

import openpyxl
from collections import Counter
import time

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book1.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()

# Step 1

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=4).value = sheet_obj.cell(row=x,column=2).value
#     for i in range(x+1,x+20):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=3).value != sheet_obj.cell(row=x,column=3).value:
#                 break
#             else:
#                 temp = sheet_obj.cell(row=x,column=4).value + ";" + sheet_obj.cell(row=i,column=2).value
#                 sheet_obj.cell(row=x,column=4).value = temp
#     x=i

# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,x+20):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=3).value != sheet_obj.cell(row=x,column=3).value:
#                 break
#             else:
#                 sheet_obj.cell(row=i,column=4).value = sheet_obj.cell(row=x,column=4).value
#     x=i

#  Step 2

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=5).value = sheet_obj.cell(row=x,column=3).value
#     for i in range(x+1,x+20):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#                 break
#             else:
#                 temp = sheet_obj.cell(row=x,column=5).value + ";" + sheet_obj.cell(row=i,column=3).value
#                 sheet_obj.cell(row=x,column=5).value = temp
#     x=i

# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,x+20):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#                 break
#             else:
#                 sheet_obj.cell(row=i,column=5).value = sheet_obj.cell(row=x,column=5).value
#     x=i

# Step 3

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=6).value = sheet_obj.cell(row=x,column=4).value
#     for i in range(x+1,x+20):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=5).value != sheet_obj.cell(row=x,column=5).value:
#                 break
#             else:
#                 temp = sheet_obj.cell(row=x,column=6).value + ";" + sheet_obj.cell(row=i,column=4).value
#                 sheet_obj.cell(row=x,column=6).value = temp
#     x=i

# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,x+20):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=5).value != sheet_obj.cell(row=x,column=5).value:
#                 break
#             else:
#                 sheet_obj.cell(row=i,column=6).value = sheet_obj.cell(row=x,column=6).value
#     x=i

# for x in range(2,sheet_obj.max_row+1):
#     peopleString = str(sheet_obj.cell(row = x, column = 6).value)
#     peopleArrayToSort = peopleString.split(';')
#     peopleArrayToSort = list(dict.fromkeys(peopleArrayToSort))
#     peopleListToStr = ';'.join(map(str,peopleArrayToSort))
#     sheet_obj.cell(row = x, column = 6).value = peopleListToStr

# Step 4

for x in range(2,sheet_obj.max_row+1):
    sheet_obj.cell(row=x,column=7).value = sheet_obj.cell(row=x,column=5).value
    for i in range(x+1,x+20):
        if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
            break
        else:
            if sheet_obj.cell(row=i,column=6).value != sheet_obj.cell(row=x,column=6).value:
                break
            else:
                temp = sheet_obj.cell(row=x,column=7).value + ";" + sheet_obj.cell(row=i,column=5).value
                sheet_obj.cell(row=x,column=7).value = temp
    x=i

for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,x+20):
        if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
            break
        else:
            if sheet_obj.cell(row=i,column=6).value != sheet_obj.cell(row=x,column=6).value:
                break
            else:
                sheet_obj.cell(row=i,column=7).value = sheet_obj.cell(row=x,column=7).value
    x=i

for x in range(2,sheet_obj.max_row+1):
    peopleString = str(sheet_obj.cell(row = x, column = 7).value)
    peopleArrayToSort = peopleString.split(';')
    peopleArrayToSort = list(dict.fromkeys(peopleArrayToSort))
    peopleListToStr = ';'.join(map(str,peopleArrayToSort))
    sheet_obj.cell(row = x, column = 7).value = peopleListToStr

wb_obj.save(path)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))