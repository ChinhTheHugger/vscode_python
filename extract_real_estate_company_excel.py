import asyncio
from playwright.async_api import async_playwright
import openpyxl
from bs4 import BeautifulSoup
import re
from docx import Document
import os
import time
import json
import pandas

new_folder_path = 'C:\\Users\\phams\\Downloads\\doanh_nghiep\\thong tin doanh nghiep'
os.makedirs(new_folder_path, exist_ok=True)

types = [
    ('chu-dau-tu',10),
    ('nha-thau-tu-van',2),
    ('tong-thau-du-an',4),
    ('thau-co-dien-vat-tu-thi-cong',1),
    ('thau-hang-muc-xay-dung-vat-tu-thi-cong',1),
    ('dich-vu-logicstic',1),
    ('cong-nghe-va-thiet-bi-san-xuat',1)
]

types_new = ['Chủ đầu tư','Nhà thầu tư vấn','Tổng thầu dự án','Thầu cơ điện vật tư thi công','Thầu hạng mục xây dựng vật tư thi công','Dịch vụ Logistic','Công nghệ và thiết bị sản xuất']

# types = [('gian-hang',2)]

df_list = []

for idx, obj in enumerate(types):
    spreadsheet_original = openpyxl.load_workbook(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\{obj[0]}.xlsx')
    sheet_original = spreadsheet_original.active
    
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active
        
    sheet.cell(row=1,column=1).value = 'Phân loại'
    sheet.cell(row=1,column=2).value = 'Tên công ty'
    sheet.cell(row=1,column=3).value = 'Loại công ty'
    sheet.cell(row=1,column=4).value = 'Địa chỉ'
    sheet.cell(row=1,column=5).value = 'Điện thoại'
    sheet.cell(row=1,column=6).value = 'Email'
    sheet.cell(row=1,column=7).value = 'Website'
    sheet.cell(row=1,column=8).value = 'Giới thiệu'
    sheet.cell(row=1,column=9).value = 'Dự án'
    sheet.cell(row=1,column=10).value = 'Sản phẩm'
    sheet.cell(row=1,column=11).value = 'Mã số doanh nghiệp'
    sheet.cell(row=1,column=12).value = 'Năm thành lập'
    sheet.cell(row=1,column=13).value = 'Số lượng nhân viên'
    sheet.cell(row=1,column=14).value = 'Hồ sơ kinh nghiệm'
    sheet.cell(row=1,column=15).value = 'Page source'
    
    def extract_between_keywords(big_string, keyword1, keyword2):
        # Find the starting position of keyword1
        start_pos = big_string.find(keyword1)
        
        # If keyword1 is not found, return an empty string
        if start_pos == -1:
            return ""
        
        # Adjust start_pos to the end of keyword1
        start_pos += len(keyword1)
        
        # Find the starting position of keyword2, starting from the end of keyword1
        end_pos = big_string.find(keyword2, start_pos)
        
        # If keyword2 is not found, return an empty string
        if end_pos == -1:
            return ""
        
        # Extract and return the substring between keyword1 and keyword2
        result_str = str(big_string[start_pos:end_pos])
        
        result_str.replace(':','')
        result_str.replace('"','')
        result_str.strip()
        
        return str(result_str)
    checkpoints = []
    
    for i in range(2,sheet_original.max_row+1):
        with open(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\page_sources\\{obj[0]}_{i-1}.html', 'r', encoding='utf-8') as file:
            html_content = file.read()
            
        # Parse the HTML content
        soup = BeautifulSoup(html_content, 'html.parser')

        checkpoint_text = None
        detail_article = soup.find('div', class_='d-flex d-md-none row')
        if detail_article:
            h1_tag = detail_article.find('h1')
            if h1_tag:
                checkpoint_text = h1_tag.get_text(separator='\n').strip()

        arr = []
        
        if checkpoint_text and checkpoint_text not in checkpoints and checkpoint_text != '':
            checkpoints.append(checkpoint_text)
            
            checkpoint_text = str(checkpoint_text).replace('/','-')
            
            # Find the div with class 'special-class'
            special_div = soup.find('div', class_='col-12 col-md-8')

            # Extract all text inside that div
            if special_div:
                for text_element in special_div.stripped_strings:
                    arr.append(text_element)
                            
            # Find the div with class 'special-class'
            special_div = soup.find('div', class_='col-12 col-md-4 mt-3')

            # Extract all text inside that div
            if special_div:
                for text_element in special_div.stripped_strings:
                    arr.append(text_element)

            sheet.cell(row=i,column=1).value = types_new[idx]
            
            if len(arr) != 0:
                sheet.cell(row=i,column=2).value = checkpoint_text
                sheet.cell(row=i,column=3).value = arr[0]
                sheet.cell(row=i,column=4).value = arr[1]
                
                arr.append('Stop text')
                string_merged = ' '.join(arr)
                
                sheet.cell(row=i,column=5).value = extract_between_keywords(string_merged,'Điện thoại:','Email:')
                
                sheet.cell(row=i,column=6).value = extract_between_keywords(string_merged,'Email:','Website:')
                
                sheet.cell(row=i,column=7).value = extract_between_keywords(string_merged,'Website:','Giới thiệu ')
                
                sheet.cell(row=i,column=8).value = extract_between_keywords(string_merged,'Giới thiệu ','Dự án ')
                
                sheet.cell(row=i,column=9).value = extract_between_keywords(string_merged,'Dự án ','Sản phẩm ')
                
                sheet.cell(row=i,column=10).value = extract_between_keywords(string_merged,'Sản phẩm ','Mã số doanh nghiệp:')
                
                sheet.cell(row=i,column=11).value = extract_between_keywords(string_merged,'Mã số doanh nghiệp:','Năm thành lập:')
                
                sheet.cell(row=i,column=12).value = extract_between_keywords(string_merged,'Năm thành lập:','Số lượng nhân viên:')
                
                sheet.cell(row=i,column=13).value = extract_between_keywords(string_merged,'Số lượng nhân viên:','Hồ sơ kinh nghiệm:')
                
                sheet.cell(row=i,column=14).value = extract_between_keywords(string_merged,'Hồ sơ kinh nghiệm:','Stop text')
                
                sheet.cell(row=i,column=15).value = i - 1
            else:
                continue

    spreadsheet.save(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\tong_hop_doanh_nghiep_{obj[0]}.xlsx')
    
    df = pandas.read_excel(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\tong_hop_doanh_nghiep_{obj[0]}.xlsx')
    
    df.drop_duplicates(subset='Tên công ty', ignore_index=True)
    df.dropna(subset='Tên công ty')
    
    df.to_excel(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\tong_hop_doanh_nghiep_{obj[0]}.xlsx', index=False)