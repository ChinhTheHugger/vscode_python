import openpyxl
from collections import Counter

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.02 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V17 V4 - Copy - Copy.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# * Check houses of each person in a shared phone group and remove unique houses of people with large number of houses to reduce string length

# def filterHouse(stringOne,stringTwo,stringBase):
#     arrayOne = stringOne.split(';')
#     arrayTwo = stringTwo.split(';')
#     arrayFilter = stringBase.split(';')
#     for ele in arrayOne:
#         if ele in arrayTwo and arrayFilter.count(ele)==0:
#             arrayFilter.append(ele)
#     arrayFiltered = list(dict.fromkeys(arrayFilter))
#     stringFiltered = ';'.join(map(str,arrayFiltered))
#     return stringFiltered

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row = x, column = 11).value = "temp"
#     if sheet_obj.cell(row = x, column = 2).value == "BIỆN XUÂN KHEN":
#         baseString = sheet_obj.cell(row = x, column = 7).value
#         count = 0
#         for y in range(x+1,x+1000):
#             if sheet_obj.cell(row = y, column = 4).value != sheet_obj.cell(row = x, column = 4).value:
#                 count = y
#                 break
#         for i in range(x+1,count):
#             sheet_obj.cell(row = x, column = 11).value = filterHouse(sheet_obj.cell(row = i, column = 7).value,baseString,sheet_obj.cell(row = x, column = 11).value)
#     x = count

# ***

# wb_obj = openpyxl.load_workbook(path)
# wb_obj.active = wb_obj.worksheets[1]
# sheet_obj = wb_obj.active
# for x in range(2,sheet_obj.max_row+1):
#     count = 0
#     for y in range(x+1,x+1000):
#         if sheet_obj.cell(row = y, column = 4).value != sheet_obj.cell(row = x, column = 4).value:
#             break
#         else:
#             if sheet_obj.cell(row = y, column = 2).value == "BIỆN XUÂN KHEN":
#                 sheet_obj.cell(row = y, column = 2).value = None
#     x = y

# ***

# * Group with the original and  filter

# def GroupAndFilter(stringOne,stringTwo,stringDestination):
#     arrayOne = stringOne.split(';')
#     arrayTwo = stringTwo.split(';')
#     arrayFilter = stringDestination.split(';')
#     for ele in arrayOne:
#         if ele in arrayTwo:
#             arrayFilter.append(ele)
#     arrayFiltered = list(dict.fromkeys(arrayFilter))
#     stringFiltered = ';'.join(map(str,arrayFiltered))
#     return stringFiltered

# for x in range(2,sheet_obj.max_row+1):
#     if sheet_obj.cell(row = x, column = 8).value == None:
#         sheet_obj.cell(row = x, column = 8).value = "temp"
#         sheet_obj.cell(row = x, column = 8).value = GroupAndFilter(sheet_obj.cell(row = x, column = 11).value,sheet_obj.cell(row = x, column = 7).value,sheet_obj.cell(row = x, column = 8).value)
#     else:
#         continue

# ***

# * Count number of house and difference

# for x in range(2,sheet_obj.max_row+1):
#     houseCountOne = sheet_obj.cell(row = x, column = 7).value
#     houseCountTwo = sheet_obj.cell(row = x, column = 8).value
#     sheet_obj.cell(row = x, column = 12).value = houseCountOne.count(';')-houseCountTwo.count(';')
    
wb_obj.save(path)