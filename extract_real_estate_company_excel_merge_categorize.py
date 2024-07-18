import asyncio
from playwright.async_api import async_playwright
import openpyxl
from bs4 import BeautifulSoup
import re
from docx import Document
import os
import time
import json
import pandas

file_path = f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\tong_hop_doanh_nghiep.xlsx'

spreadsheet = openpyxl.load_workbook(file_path)
spreadsheet.create_sheet('Sheet2')
sheet1 = spreadsheet['Sheet1']
sheet2 = spreadsheet['Sheet2']

types = []

# for i in range(2,sheet1.max_row+1):
#     string = str(sheet1.cell(row=i,column=4).value)
#     array = string.split(',')
#     for substr in array:
#         if substr.strip() not in types:
#             types.append(substr.strip())

# columns = ['Phân loại','Tên công ty']

# for val in types:
#     columns.append(val)
    
# print(columns)

columns = ['Phân loại', 'Tên công ty', 'Chủ đầu tư', 'CĐT dự án vốn trong nước', 'CĐT dự án vốn FDI', 'Tổng thầu dự án', 'Tổng thầu xây dựng', 'Nhà thầu tư vấn', 'Tư vấn giám sát', 'Tư vấn quản lý dự án', 
          'CĐT Khu đô thị - Khu CN', 'Tổng thầu thiết bị', 'Tư vấn đầu tư', 'Tư vấn thiết kế dự án', 'Thầu cơ điện ( Vật tư, thi công)', 'Thầu hạng mục xây dựng (vật tư, thi công)', 
          'Kết cấu khung phần thô & Xây dựng', 'Thầu khác', 'Dịch vụ & Logicstic', 'Vận tải - Locgic stic', 'Nhà thầu thiết bị cơ điện', 'Công nghệ và thiết bị sản xuất', 'Thiết bị phục vụ sản xuất', 
          'Nội thất', 'Tổng thầu cơ điện', 'Dịch vụ pháp lý', 'tài chính', 'kiểm toán', 'bảo hiểm', 'Thầu HVAC', 'Nền móng & Phần ngầm', 'Kết cấu thép', 'Thi công hoàn thiện', 'Thầu âm thanh thiết bị camera', 
          'hệ thống ASM', 'mạng viễn thông', 'Thầu cấp thoát nước. xử lý nước thải', 'Thầu khí', 'phòng sạch', 'kho lạnh', 'Thầu phòng cháy chữa cháy', 'Thầu thang máy', 'cần trục', 'thiết bị nâng hạ', 
          'cân điện tử', 'Nhà thầu vật liệu giao thông', 'San lấp mặt bằng & Hạ tầng', 'cảnh quan', 'Cải tạo và sửa chữa', 'Cung cấp nhân sự', 'phiên dịch & đưa đón chuyên gia', 
          'Vệ sinh công nghiệp - Suất ăn công nghiệp', 'Thiết bị phục vụ thi công', 'Tự động hóa dây chuyền sản xuất', 'tự động hóa', 'Máy phát điện - Máy động cơ']

df = pandas.DataFrame(columns=columns)

data = []

for i in range(2,sheet1.max_row+1):
    row = {}
    row['Phân loại'] = sheet1.cell(row=i,column=2).value
    row['Tên công ty'] = sheet1.cell(row=i,column=3).value
    for val in types:
        if val in str(sheet1.cell(row=i,column=4).value):
            row[val] = 'X'
        else:
            row[val] = ''
    data.append(row)

# df = df._append(data)

# book = openpyxl.load_workbook(file_path)
# try:
#     with pandas.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         df.to_excel(writer, sheet_name='Sheet2', index=True)
# except FileNotFoundError:
#     with pandas.ExcelWriter(file_path, engine='openpyxl') as writer:
#         df.to_excel(writer, sheet_name='Sheet2', index=True)

# column_widths_in_inches = {'A':0.5,'B':2,'C':3.25}

# for col, width in column_widths_in_inches.items():
#     # Convert inches to Excel column width units
#     excel_column_width = width * 10.71
#     sheet2.column_dimensions[col].width = excel_column_width

# for idx, col in enumerate(sheet2.iter_cols()):
#     if idx > 3:
#         sheet2.column_dimensions[str(col[0].value)].width = 1.5 * 10.71
#     else:
#         continue

# spreadsheet.save(file_path)