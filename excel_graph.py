import openpyxl
import pandas
from collections import Counter
from difflib import SequenceMatcher
from collections import OrderedDict
import time
import numpy
import igraph

import sys

pathTest = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book1.XLSX"
path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.18 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

wb_test = openpyxl.load_workbook(pathTest)
sh_test = wb_test.active

start = time.time()
print("Start time = " + time.strftime("%H:%M:%S", time.gmtime(start)))
time.sleep(5)



def phGraph(startNum,endNum):
    phArray = []
    for x in range(startNum,endNum):
        if str(sheet_obj.cell(row=x,column=3).value) not in phArray:
            phArray.append(str(sheet_obj.cell(row=x,column=3).value))
        if str(sheet_obj.cell(row=x,column=4).value) not in phArray:
            phArray.append(str(sheet_obj.cell(row=x,column=4).value))
    phArray = sorted(phArray)
    for a in range(startNum,endNum):
        sheet_obj.cell(row=a,column=5).value = phArray.index(str(sheet_obj.cell(row=a,column=3).value))
        sheet_obj.cell(row=a,column=6).value = phArray.index(str(sheet_obj.cell(row=a,column=4).value))
    phGraph = igraph.Graph(n=len(phArray))
    phGraph.vs['name'] = phArray
    for x in range(startNum,endNum):
        phGraph.add_edges([(sheet_obj.cell(row=x,column=5).value,sheet_obj.cell(row=x,column=6).value)])
    for s in phGraph.components().subgraphs():
        pArr = []
        hArr = []
        for ele in s.vs['name']:
            if "-" in ele:
                hArr.append(ele)
            else:
                pArr.append(ele)
        for x in range(startNum,endNum):
            if str(sheet_obj.cell(row=x,column=3).value) in pArr:
                sheet_obj.cell(row=x,column=7).value = ';'.join(sorted(pArr))
                sheet_obj.cell(row=x,column=8).value = ';'.join(sorted(hArr))
    return

arr = [2]
for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,sheet_obj.max_row+2):
        if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
            break
    if i not in arr:
        arr.append(i)
    x=i

# print(arr)
print(len(arr))

newArr = numpy.array(arr)
for x in range(0,len(newArr)-1):
    if newArr[x+1]-newArr[x]==1:
        sheet_obj.cell(row=newArr[x],column=7).value = str(sheet_obj.cell(row=newArr[x],column=3).value)
        sheet_obj.cell(row=newArr[x],column=8).value = str(sheet_obj.cell(row=newArr[x],column=4).value)
    else:
        for i in range(newArr[x],newArr[x+1]):
            phGraph(newArr[x],newArr[x+1])

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=9).value = str(sheet_obj.cell(row=x,column=7).value).count(';') + 1
#     sheet_obj.cell(row=x,column=10).value = str(sheet_obj.cell(row=x,column=8).value).count(';') + 1



wb_obj.save(path)

end = time.time()

print("Run time = " + time.strftime("%H:%M:%S", time.gmtime(end-start)))
print("End time = " + time.strftime("%H:%M:%S", time.gmtime(end)))