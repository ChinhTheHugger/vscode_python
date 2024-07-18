import asyncio
from playwright.async_api import async_playwright
import openpyxl
from bs4 import BeautifulSoup
import re
from docx import Document
import os
import time
import json

types = [
    ('chu-dau-tu',10),
    ('nha-thau-tu-van',2),
    ('tong-thau-du-an',4),
    ('thau-co-dien-vat-tu-thi-cong',1),
    ('thau-hang-muc-xay-dung-vat-tu-thi-cong',1),
    ('dich-vu-logicstic',1),
    ('cong-nghe-va-thiet-bi-san-xuat',1)
]

# types = [('gian-hang',2)]

for obj in types:
    spreadsheet = openpyxl.load_workbook(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\{obj[0]} - Copy.xlsx')
    sheet = spreadsheet.active
    
    for i in range(2,sheet.max_row+1):
        base_link = sheet.cell(row=i,column=2).value
        
        if base_link == '':
            continue
        else:
            for x in range(i+1,i+5):
                if sheet.cell(row=x,column=2).value == base_link:
                    sheet.cell(row=x,column=2).value = None
    
    spreadsheet.save(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\{obj[0]} - Copy.xlsx')
    
    print(f'Cleared all duplicate links in {obj[0]} - Copy.xlsx')