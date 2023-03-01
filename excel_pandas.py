import pandas as pd
import openpyxl

# ***

# * Test pandas

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.01 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V17 V4 - Copy.XLSX"
# main_sheet = pd.read_excel(path,sheet_name='Sheet1')
# process_sheet = pd.read_excel(path,sheet_name='Sheet2')
# print(process_sheet)
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
count=0
# for x in range(2,sheet.max_row+1):
#     for y in range(x+1,sheet.max_row+1):
#         if sheet.cell(row=y,column=4).value != sheet.cell(row=x,column=4).value:
#             count=y
#             break
#     for i in range(x+1,count):
#         if sheet.cell(row=i,column=4).value == sheet.cell(row=x,column=4).value and sheet.cell(row=i,column=2).value != sheet.cell(row=x,column=2).value:
#             sheet.cell(row=i,column=3).value=None
#     x=count
for x in range(2,sheet.max_row+1):
    if sheet.cell(row=x,column=2).value == "BIỆN XUÂN KHEN":
        phone=sheet.cell(row=x,column=9).value
        phoneList=phone.split(';')
        for ph in phoneList:
            for i in range(2,sheet.max_row+1):
                if sheet.cell(row=i,column=3).value==ph:
                    sheet.cell(row=i,column=3).value=None
workbook.save(path)