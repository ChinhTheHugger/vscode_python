import asyncio
from playwright.async_api import async_playwright
import openpyxl
from bs4 import BeautifulSoup
import re
from docx import Document
import os
import time
import json

# types = [
#     ('chu-dau-tu',10),
#     ('nha-thau-tu-van',2),
#     ('tong-thau-du-an',4),
#     ('thau-co-dien-vat-tu-thi-cong',1),
#     ('thau-hang-muc-xay-dung-vat-tu-thi-cong',1),
#     ('dich-vu-logicstic',1),
#     ('cong-nghe-va-thiet-bi-san-xuat',1)
# ]

types = [('gian-hang',2)]

urls = []

new_folder_path = 'C:\\Users\\phams\\Downloads\\doanh_nghiep'
os.makedirs(new_folder_path, exist_ok=True)

for idx, obj in enumerate(types):
    name, count = obj
    for page in range(count):
        # url = f'https://fcivietnam.com/thong-tin-doanh-nghiep/{name}?size=100&p={page+1}'
        url = f'https://fcivietnam.com/thong-tin-doanh-nghiep/{name}?p={page+1}'
        if url not in urls:
            urls.append((url,idx))
        
async def get_accessibility_tree(url_list):
    async with async_playwright() as p:
        browser = await p.firefox.launch_persistent_context(user_data_dir="C:\\Users\\phams\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\bd5r7ma0.default-release - Copy", headless=True)
        page = await browser.new_page()
                
        for idx, url in enumerate(url_list):
            if url[1] == 0:
                await page.goto(url[0])
                
                com_type = types[url[1]][0]
                        
                # Get the accessibility tree
                accessibility_snapshot = await page.accessibility.snapshot()
                        
                with open(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\{com_type}_{idx+1}.json', 'w') as f:
                    json.dump(accessibility_snapshot, f, indent=4)
                        
                print(f'{com_type}: saved result page {idx+1}')
            else:
                await page.goto(url[0])
                
                sum = 0
                for num in range(url[1]):
                    sum += types[num][1]
                
                com_type = types[url[1]][0]
                        
                # Get the accessibility tree
                accessibility_snapshot = await page.accessibility.snapshot()
                        
                with open(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\{com_type}_{idx+1-sum}.json', 'w') as f:
                    json.dump(accessibility_snapshot, f, indent=4)
                        
                print(f'{com_type}: saved result page {idx+1-sum}')
                    
        await browser.close()
        
asyncio.get_event_loop().run_until_complete(get_accessibility_tree(urls))


        
        # file_path = f'C:\\Users\\phams\\Downloads\\{name}.json'

        # new_folder_path = 'C:\\Users\\phams\\Downloads\\doanh_nghiep'
        # os.makedirs(new_folder_path, exist_ok=True)

        # spreadsheet = openpyxl.Workbook()
        # sheet = spreadsheet.active

        # sheet.cell(row=1,column=1).value = 'Page source'
        # sheet.cell(row=1,column=2).value = 'Link'
        # sheet.cell(row=1,column=3).value = 'Name'
        # sheet.cell(row=1,column=4).value = 'Type'

        # try:
        #     with open(file_path, 'r', encoding='utf-8') as file:
        #         data = json.load(file)
        # except UnicodeDecodeError:
        #     # If 'utf-8' encoding fails, try 'iso-8859-1'
        #     with open(file_path, 'r', encoding='iso-8859-1') as file:
        #         data = json.load(file)
                
        # links = []

        # # print(len(data['children'][2]['children'][13]['children']))

        # for idx, com in enumerate(data['children'][2]['children'][13]['children']):
        #     if com['value'] not in links:
        #         links.append(com['value'])
                
        #         sheet.cell(row=idx+2,column=1).value = idx + 1
        #         sheet.cell(row=idx+2,column=2).value = com['value']

        # print(len(links))