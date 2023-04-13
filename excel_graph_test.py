import openpyxl
import pandas
from collections import Counter
from difflib import SequenceMatcher
from collections import OrderedDict
import time
import numpy
import igraph

import sys

pathTest = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book1.XLSX"
path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.13 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_base = wb_obj.active

wb_test = openpyxl.load_workbook(pathTest)
sh_test = wb_test.active

start = time.time()
print("Start time = " + time.strftime("%H:%M:%S", time.gmtime(start)))
time.sleep(5)

graph = igraph.Graph(n=8,edges=[(0,4),(1,4),(1,6),(2,5),(3,6),(3,7)])
graph.vs['name'] = ['H1','H2','H3','H4','P1','P2','P3','P4']

print(graph)

# for s in graph.components().subgraphs():
#     print(s.vs['name'])
#     print(' | '.join(s.vs['name']))
#     hStr = []
#     pStr = []
#     for ele in s.vs['name']:
#         if "H" in ele:
#             hStr.append(ele)
#         if "P" in ele:
#             pStr.append(ele)
#     print(';'.join(hStr) + " and " + ';'.join(pStr))

# n = 12
# graphTest = igraph.Graph(n)
# for x in range(0,5):
#     graphTest.add_edges([(x,n-1-x)])
# print(graphTest)

phList = ['H1','H2','H3','H4','P1','P2','P3','P4']
graphTest = igraph.Graph(n=len(phList))
graphTest.vs['name'] = phList
for x in range(2,sh_test.max_row+1):
    # for idx,val in enumerate(graphTest.vs['name']):
    #     if sh_test.cell(row=x,column=2).value == val:
    #         sh_test.cell(row=x,column=4).value = idx
    #     elif sh_test.cell(row=x,column=3).value == val:
    #         sh_test.cell(row=x,column=5).value = idx
    sh_test.cell(row=x,column=4).value = phList.index(sh_test.cell(row=x,column=2).value)
    sh_test.cell(row=x,column=5).value = phList.index(sh_test.cell(row=x,column=3).value)
for i in range(2,sh_test.max_row+1):
    graphTest.add_edges([(sh_test.cell(row=i,column=4).value,sh_test.cell(row=i,column=5).value)])
print(graphTest)

wb_test.save(pathTest)

end = time.time()

print("Run time = " + time.strftime("%H:%M:%S", time.gmtime(end-start)))