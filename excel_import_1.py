import openpyxl
from difflib import SequenceMatcher
from collections import OrderedDict
#import xlsxwriter

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.01 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V17 V4 - Copy - Copy.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# * Select a portion of code and Ctrl+/ to comment the entire selection

# ***

# * Test import, edit and save

# print(sheet_obj.max_row)
# cell_obj = sheet_obj.cell(row = 1, column = 1)
# cell_obj.value = "Count (written in python)"
# wb_obj.save("E:\Pham Thanh Quyet - 23.12.2022\DSKH 22.12.23\VRS VRH\Book1.xlsx")
# print(cell_obj.value)

# ***

# * Calculate similarity between names

# cell_obj = sheet_obj.cell(row = 1, column = 4)
# cell_obj.value = "Similarity in percentage (upward)"
# for x in range(3,sheet_obj.max_row+1):
#    s = SequenceMatcher(None, sheet_obj.cell(row = x-1, column = 2).value, sheet_obj.cell(row = x, column = 2).value)
#    cell_obj = sheet_obj.cell(row = x, column = 4)
#    cell_obj.value = s.ratio()
# wb_obj.save("E:\Pham Thanh Quyet - 23.12.2022\DSKH 22.12.23\VRS VRH\Book1.xlsx")

# ***

# * Mark high percentage

# for x in range(2,sheet_obj.max_row):
#    if sheet_obj.cell(row = x, column = 3).value >= 0.91:
#        sheet_obj.cell(row = x, column = 6).value = "x"
#        sheet_obj.cell(row = x+1, column = 6).value = "x"
# wb_obj.save("E:\Pham Thanh Quyet - 23.12.2022\DSKH 22.12.23\VRS VRH\Book1.xlsx")

# ***

# * Transpose house by phone

# for x in range(2,sheet_obj.max_row+1):
#    sheet_obj.cell(row = x, column = 9).value = sheet_obj.cell(row = x, column = 7).value
#    for i in range(x+1,x+5500):
#        if sheet_obj.cell(row = i, column = 2).value == sheet_obj.cell(row = x, column = 2).value and sheet_obj.cell(row = i, column = 4).value == sheet_obj.cell(row = x, column = 4).value:
#            sheet_obj.cell(row = x, column = 9).value = sheet_obj.cell(row = x, column = 9).value + ";" + sheet_obj.cell(row = i, column = 7).value
#            sheet_obj.cell(row = i, column = 3).value = None
#            continue
#        else:
#            break
#    x = i

# ***

# * Transpose phone, house group by house

# XlsxWriter: workbook = xlsxwriter.Workbook(path)
# XlsxWriter: worksheet = workbook.get
# for x in range(2,sheet_obj.max_row+1):
#    #phone
#    sheet_obj.cell(row = x, column = 9).number_format = '@' #change data type in cell to string
#    sheet_obj.cell(row = x, column = 9).value = sheet_obj.cell(row = x, column = 4).value
#    #XlsxWriter: wb_obj.write_string(x,10,sheet_obj.cell(row = x, column = 4).value)
#    #house group
#    sheet_obj.cell(row = x, column = 10).number_format = '@' #change data type in cell to string
#    sheet_obj.cell(row = x, column = 10).value = sheet_obj.cell(row = x, column = 8).value
#    #XlsxWriter: wb_obj.write_string(x,12,sheet_obj.cell(row = x, column = 8).value)
#    for i in range(x+1,x+5500):
#         if sheet_obj.cell(row = i, column = 2).value == sheet_obj.cell(row = x, column = 2).value and sheet_obj.cell(row = i, column = 7).value == sheet_obj.cell(row = x, column = 7).value:
#            #phone
#            sheet_obj.cell(row = x, column = 9).value = sheet_obj.cell(row = x, column = 9).value + ";" + sheet_obj.cell(row = i, column = 4).value
#            #XlsxWriter: wb_obj.write_string(x,10,sheet_obj.cell(row = x, column = 10).value + ";" + sheet_obj.cell(row = i, column = 4).value)
#            #house group
#            sheet_obj.cell(row = x, column = 10).value = sheet_obj.cell(row = x, column = 10).value + ";" + sheet_obj.cell(row = i, column = 8).value
#            #XlsxWriter: wb_obj.write_string(x,12,sheet_obj.cell(row = x, column = 12).value + ";" + sheet_obj.cell(row = i, column = 8).value)
#            sheet_obj.cell(row = i, column = 3).value = None
#            continue
#         else:
#            break
#    x = i

# ***

# * Remove duplicate and sort

# for x in range(2,sheet_obj.max_row+1):
#    #phone
#    phoneString = sheet_obj.cell(row = x, column = 9).value
#    phoneArrayToSort = phoneString.split(';')
#    phoneArrayToSort = list(set(phoneArrayToSort))
#    phoneListToStr = ';'.join(map(str,phoneArrayToSort))
#    sheet_obj.cell(row = x, column = 9).value = phoneListToStr
#    #house
#    houseString = sheet_obj.cell(row = x, column = 10).value
#    houseArrayToSort = houseString.split(';')
#    houseArrayToSort = list(set(houseArrayToSort))
#    houseListToStr = ';'.join(map(str,houseArrayToSort))
#    sheet_obj.cell(row = x, column = 10).value = houseListToStr

# ***

# * Group people sharing phone number

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row = x, column = 12).number_format = '@'
#     sheet_obj.cell(row = x, column = 12).value = sheet_obj.cell(row = x, column = 11).value
#     for i in range(x+1,x+5000):
#         if sheet_obj.cell(row = i, column = 4).value == sheet_obj.cell(row = x, column = 4).value:
#             sheet_obj.cell(row = x, column = 12).value = sheet_obj.cell(row = x, column = 12).value + "&" + sheet_obj.cell(row = i, column = 11).value
#             sheet_obj.cell(row = i, column = 3).value = None
#             continue
#         else:
#             break
#     x = i

# ***

# * Remove duplicate and sort people

# for x in range(2,sheet_obj.max_row+1):
#     peopleString = sheet_obj.cell(row = x, column = 12).value
#     peopleArrayToSort = peopleString.split('&')
#     peopleArrayToSort = list(dict.fromkeys(peopleArrayToSort))
#     peopleListToStr = '&'.join(map(str,peopleArrayToSort))
#     sheet_obj.cell(row = x, column = 12).value = peopleListToStr

# ***

# * Check houses of each person in a shared phone group and remove unique houses of people with large number of houses to reduce string length

# for x in range(2,sheet_obj.max_row+1):
#     waymark = 0
#     if sheet_obj.cell(row = x, column = 2).value == "BIỆN XUÂN KHEN":
#         waymark = x
#     for i in range(waymark+1,2000):
#         if sheet_obj.cell(row = i, column = 2).value == "BIỆN XUÂN KHEN":
#             sheet_obj.cell(row = i, column = 14).value = "x"    
wb_obj.save(path)