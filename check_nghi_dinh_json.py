import asyncio
from playwright.async_api import async_playwright
import json
from datetime import date
import openpyxl
import os

path = 'C:\\Users\\phams\\Downloads\\nghi dinh\\nghi_dinh.xlsx'

if not os.path.exists(path):
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active
    
    sheet.cell(row=1,column=1).value = 'Num.'
    sheet.cell(row=1,column=2).value = 'Date'
    sheet.cell(row=1,column=3).value = 'JSON'
    sheet.cell(row=1,column=4).value = 'Sheet'
    sheet.cell(row=1,column=5).value = 'File name'
    
    spreadsheet.save(path)

spreadsheet = openpyxl.load_workbook(path)
sheet = spreadsheet.active

async def get_accessibility_tree(url):
    async with async_playwright() as p:
        browser = await p.firefox.launch_persistent_context(user_data_dir="C:\\Users\\phams\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\bd5r7ma0.default-release - Copy", headless=True)
        page = await browser.new_page()
        await page.goto(url)
        
        # Get the accessibility tree
        accessibility_snapshot = await page.accessibility.snapshot()
        
        today = date.today()
        
        with open(f'C:\\Users\\phams\\Downloads\\nghi dinh\\json\\nghi_dinh_list_{today}.json', 'w') as f:
            json.dump(accessibility_snapshot, f, indent=2)
        
        rows = sheet.max_row
        
        if sheet.cell(row=rows,column=3).value != f'json\\nghi_dinh_list_{today}.json':
            sheet.cell(row=rows+1,column=1).value = rows
            sheet.cell(row=rows+1,column=2).value = today
            sheet.cell(row=rows+1,column=3).value = f'json\\nghi_dinh_list_{today}.json'
            sheet.cell(row=rows+1,column=4).value = f'sheet\\nghi_dinh_list_{today}.xlsx'
        
        print(f'Decree lists, date: {today}')
        
        await browser.close()
        
        # # Get the page source
        # page_source = await page.content()
        
        # # Save the page source to an HTML file
        # with open('C:\\Users\\phams\\Downloads\\page_source.html', 'w', encoding='utf-8') as f:
        #     f.write(page_source)
        
        # await browser.close()

# URL to open
url = 'https://danhmuchanhchinh.gso.gov.vn/NghiDinh.aspx'

# Run the function
asyncio.get_event_loop().run_until_complete(get_accessibility_tree(url))

spreadsheet.save(path)

