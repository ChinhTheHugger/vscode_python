import matplotlib.pyplot as plt
import openpyxl
import io

spreadsheet = openpyxl.load_workbook("C:\\Users\\phams\\Downloads\\linestring_analysis.xlsx")
sheet = spreadsheet.active

car_polate = []
car_buffer = []

bike_polate = []
bike_buffer = []

foot_polate = []
foor_buffer = []

motor_polate = []
motor_buffer = []

for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i,column=7).value != 0:
        # car_polate.append((sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=5).value))
        car_buffer.append((sheet.cell(row=i,column=11).value,sheet.cell(row=i,column=15).value))
    if sheet.cell(row=i,column=8).value != 0:
        # bike_polate.append((sheet.cell(row=i,column=3).value,sheet.cell(row=i,column=7).value))
        bike_buffer.append((sheet.cell(row=i,column=12).value,sheet.cell(row=i,column=16).value))
    if sheet.cell(row=i,column=9).value != 0:
        # foot_polate.append((sheet.cell(row=i,column=3).value,sheet.cell(row=i,column=9).value))
        foor_buffer.append((sheet.cell(row=i,column=13).value,sheet.cell(row=i,column=17).value))
    if sheet.cell(row=i,column=10).value != 0:
        # motor_polate.append((sheet.cell(row=i,column=3).value,sheet.cell(row=i,column=9).value))
        motor_buffer.append((sheet.cell(row=i,column=14).value,sheet.cell(row=i,column=18).value))

if len(car_buffer) == 0:
    car_buffer.append((0,0))
if len(bike_buffer) == 0:
    bike_buffer.append((0,0))
if len(foor_buffer) == 0:
    foor_buffer.append((0,0))
if len(motor_buffer) == 0:
    motor_buffer.append((0,0))

car_count_p = 0
car_count_b = 0

bike_count_p = 0
bike_count_b = 0

foot_count_p = 0
foot_count_b = 0

motor_count_p = 0
motor_count_b = 0

# for c in car_polate:
#     if c[0] >= -10 and c[0] <= 10 and c[1] >= 80 and c[0] <= 100:
#         car_count_p += 1
for c in car_buffer:
    if c[0] >= -10 and c[0] <= 10 and c[1] >= 80 and c[0] <= 100:
        car_count_b += 1

# for b in bike_polate:
#     if b[0] >= -10 and b[0] <= 10 and b[1] >= 80 and b[0] <= 100:
#         bike_count_p += 1
for b in bike_buffer:
    if b[0] >= -10 and b[0] <= 10 and b[1] >= 80 and b[0] <= 100:
        bike_count_b += 1

# for f in foot_polate:
#     if f[0] >= -10 and f[0] <= 10 and f[1] >= 80 and f[0] <= 100:
#         foot_count_p += 1
for f in foor_buffer:
    if f[0] >= -10 and f[0] <= 10 and f[1] >= 80 and f[0] <= 100:
        foot_count_b += 1

# for f in motor_polate:
#     if f[0] >= -10 and f[0] <= 10 and f[1] >= 80 and f[0] <= 100:
#         motor_count_p += 1
for f in motor_buffer:
    if f[0] >= -10 and f[0] <= 10 and f[1] >= 80 and f[0] <= 100:
        motor_count_b += 1

# print(f"Num. of car route with accuracy in (-10%,10%) and overlap in (80%,100%): polate = {car_count_p}/{len(car_polate)}, buffer={car_count_b}/{len(car_buffer)}")
# print(f"Num. of bike route with accuracy in (-10%,10%) and overlap in (80%,100%): polate = {bike_count_p}/{len(bike_polate)}, buffer={bike_count_b}/{len(bike_buffer)}")
# print(f"Num. of foot route with accuracy in (-10%,10%) and overlap in (80%,100%): polate = {foot_count_p}/{len(foot_polate)}, buffer={foot_count_b}/{len(foor_buffer)}")

car = {
    'polylines': [car_buffer],
    'name': ['Car Buffer']
}

bike = {
    'polylines': [bike_buffer],
    'name': ['Bike Buffer']
}

foot = {
    'polylines': [foor_buffer],
    'name': ['Motor Buffer']
}

motor = {
    'polylines': [motor_buffer],
    'name': ['XTeam motor Buffer']
}

fig, (ax3) = plt.subplots(1, 1, figsize=(30, 12))

# for c in range(1):
#     x, y = zip(*car['polylines'][c])
#     ax1.scatter(x,y, label=car['name'][c], color='red')
# ax1.set_xlabel("Length percentage")
# ax1.set_ylabel("Similarity percentage")
# ax1.legend()
# ax1.grid(True)

# for b in range(1):
#     x, y = zip(*bike['polylines'][b])
#     ax2.scatter(x,y, label=bike['name'][b], color='blue')
# ax2.set_xlabel("Length percentage")
# ax2.set_ylabel("Similarity percentage")
# ax2.legend()
# ax2.grid(True)

for f in range(1):
    x, y = zip(*foot['polylines'][f])
    ax3.scatter(x,y, label=foot['name'][f], color='green')
ax3.set_xlabel("Length percentage")
ax3.set_ylabel("Similarity percentage")
ax3.legend()
ax3.grid(True)

# for f in range(1):
#     x, y = zip(*motor['polylines'][f])
#     ax4.scatter(x,y, label=motor['name'][f], color='orange')
# ax4.set_xlabel("Length percentage")
# ax4.set_ylabel("Similarity percentage")
# ax4.legend()
# ax4.grid(True)

img_stream = io.BytesIO()

plt.tight_layout()

plt.savefig(img_stream, format='png')
plot_file = "C:\\Users\\phams\\Downloads\\interpolate_buffer_comparison.png"
fig.savefig(plot_file)

manager = plt.get_current_fig_manager()
manager.window.state('zoomed')

plt.show()

plt.close(fig)  # Close the plot



# NOTES:
# buffer method give better result