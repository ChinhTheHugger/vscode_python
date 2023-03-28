import openpyxl
from collections import Counter
import time
import numpy

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.20 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()

# Sort

# for x in range(2,sheet_obj.max_row+1):
#     arrOne=str(sheet_obj.cell(row=x,column=5).value).split(';')
#     arrOne.sort()
#     newStrOne=';'.join(map(str,arrOne))
#     sheet_obj.cell(row=x,column=15).value=newStrOne

#     arrTwo=str(sheet_obj.cell(row=x,column=11).value).split(';')
#     arrTwo.sort()
#     newStrTwo=';'.join(map(str,arrTwo))
#     sheet_obj.cell(row=x,column=12).value=newStrTwo

#     arrThree=str(sheet_obj.cell(row=x,column=13).value).split(';')
#     arrThree.sort()
#     newStrThree=';'.join(map(str,arrThree))
#     sheet_obj.cell(row=x,column=14).value=newStrThree

# Count length

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=16).value=str(sheet_obj.cell(row=x,column=15).value).count(';')+1

# Group name with house list

# for x in range(2,sheet_obj.max_row+1):
#     if sheet_obj.cell(row=x,column=16).value>99:
#         sheet_obj.cell(row=x,column=17).value=str(sheet_obj.cell(row=x,column=2).value)+";"+str(sheet_obj.cell(row=x,column=16).value)+" căn"
#     else:
#         sheet_obj.cell(row=x,column=17).value=str(sheet_obj.cell(row=x,column=2).value)+";"+str(sheet_obj.cell(row=x,column=15).value)

# Group people sharing phone

def cleanPpl(arr):
    tmpStr='&'.join(arr)
    tmpArr=tmpStr.split('&')
    tmpArr=list(dict.fromkeys(tmpArr))
    tmpStr='&'.join(tmpArr)
    return tmpStr

def pplGroup(x,i):
    if i-x==0:
        sheet_obj.cell(row=x,column=18).value=sheet_obj.cell(row=x,column=17).value
    else:
        for a in range(x,i+1):
            pplArr=[]
            pplArr.append(sheet_obj.cell(row=a,column=17).value)
            for b in range(x,i+1):
                if sheet_obj.cell(row=b,column=17).value not in pplArr:
                    pplArr.append(str(sheet_obj.cell(row=b,column=17).value))
            pplStr=cleanPpl(pplArr)
            sheet_obj.cell(row=a,column=18).value=pplStr
    return

arr = [2]
for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,sheet_obj.max_row+2):
        if sheet_obj.cell(row=i,column=3).value != sheet_obj.cell(row=x,column=3).value:
            break
    if i not in arr:
        arr.append(i)
    x=i

print(arr)
print(len(arr))

newArr = numpy.array(arr)
for x in range(0,len(arr)-1):
    pplGroup(newArr[x],newArr[x+1]-1)



wb_obj.save(path)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))