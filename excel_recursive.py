import openpyxl
import pandas
from collections import Counter
from difflib import SequenceMatcher
from collections import OrderedDict
import time
import numpy

import sys

pathTest = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book.XLSX"
path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.08 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V23 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()



# sys.setrecursionlimit(4850)

# def cleanStr(tmpStr):
#     tmpArr=tmpStr.split(';')
#     tmpArr=list(dict.fromkeys(tmpArr))
#     tmpArr.sort()
#     tmpStrN = ';'.join(tmpArr)
#     return tmpStrN

# def phRecursion(startNum,endNum,baseNum,basePhoneStr,baseHouseStr):
#     if endNum-startNum==0:
#         if str(sheet_obj.cell(row=endNum,column=3).value) in basePhoneStr or str(sheet_obj.cell(row=endNum,column=4).value) in baseHouseStr:
#             basePhoneStr = basePhoneStr + ";" + str(sheet_obj.cell(row=endNum,column=3).value)
#             newBasePhoneStr = cleanStr(basePhoneStr)
#             baseHouseStr = baseHouseStr + ";" + str(sheet_obj.cell(row=endNum,column=4).value)
#             newBaseHouseStr = cleanStr(baseHouseStr)
#             sheet_obj.cell(row=baseNum,column=6).value = newBaseHouseStr
#             sheet_obj.cell(row=baseNum,column=5).value = newBasePhoneStr
#             return
#         else:
#             sheet_obj.cell(row=baseNum,column=6).value = cleanStr(baseHouseStr)
#             sheet_obj.cell(row=baseNum,column=5).value = cleanStr(basePhoneStr)
#             return
#     else:
#         if str(sheet_obj.cell(row=endNum,column=3).value) in basePhoneStr or str(sheet_obj.cell(row=endNum,column=4).value) in baseHouseStr:
#             basePhoneStr = basePhoneStr + ";" + str(sheet_obj.cell(row=endNum,column=3).value)
#             newBasePhoneStr = cleanStr(basePhoneStr)
#             baseHouseStr = baseHouseStr + ";" + str(sheet_obj.cell(row=endNum,column=4).value)
#             newBaseHouseStr = cleanStr(baseHouseStr)
#             return phRecursion(startNum,endNum-1,baseNum,newBasePhoneStr,newBaseHouseStr)
#         else:
#             return phRecursion(startNum,endNum-1,baseNum,basePhoneStr,baseHouseStr)



# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# # print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(newArr)-1):
#     if newArr[x+1]-newArr[x]==1:
#         sheet_obj.cell(row=newArr[x],column=5).value = str(sheet_obj.cell(row=newArr[x],column=3).value)
#         sheet_obj.cell(row=newArr[x],column=6).value = str(sheet_obj.cell(row=newArr[x],column=4).value)
#     else:
#         for i in range(newArr[x],newArr[x+1]):
#             phRecursion(newArr[x],newArr[x+1]-1,i,str(sheet_obj.cell(row=i,column=3).value),str(sheet_obj.cell(row=i,column=4).value))



sys.setrecursionlimit(4850)

def cleanStr(tmpStr):
    tmpArr=tmpStr.split(';')
    tmpArr=list(dict.fromkeys(tmpArr))
    tmpArr.sort()
    tmpStrN = ';'.join(tmpArr)
    return tmpStrN

def phRecursion(startNum,endNum,baseNum,basePhoneStr,baseHouseStr):
    phoneArr = numpy.array(str(sheet_obj.cell(row=endNum,column=5).value).split(';'))
    houseArr = numpy.array(str(sheet_obj.cell(row=endNum,column=6).value).split(';'))
    if endNum-startNum==0:
        if sum([1 for i in range(0,len(phoneArr)-1) if phoneArr[i] in basePhoneStr]) > 0 or sum([1 for i in range(0,len(houseArr)-1) if houseArr[i] in baseHouseStr]) > 0:
            basePhoneStr = basePhoneStr + ";" + str(sheet_obj.cell(row=endNum,column=5).value)
            newBasePhoneStr = cleanStr(basePhoneStr)
            baseHouseStr = baseHouseStr + ";" + str(sheet_obj.cell(row=endNum,column=6).value)
            newBaseHouseStr = cleanStr(baseHouseStr)
            sheet_obj.cell(row=baseNum,column=8).value = newBaseHouseStr
            sheet_obj.cell(row=baseNum,column=7).value = newBasePhoneStr
            return
        else:
            sheet_obj.cell(row=baseNum,column=8).value = cleanStr(baseHouseStr)
            sheet_obj.cell(row=baseNum,column=7).value = cleanStr(basePhoneStr)
            return
    else:
        if sum([1 for i in range(0,len(phoneArr)-1) if phoneArr[i] in basePhoneStr]) > 0 or sum([1 for i in range(0,len(houseArr)-1) if houseArr[i] in baseHouseStr]) > 0:
            basePhoneStr = basePhoneStr + ";" + str(sheet_obj.cell(row=endNum,column=5).value)
            newBasePhoneStr = cleanStr(basePhoneStr)
            baseHouseStr = baseHouseStr + ";" + str(sheet_obj.cell(row=endNum,column=6).value)
            newBaseHouseStr = cleanStr(baseHouseStr)
            return phRecursion(startNum,endNum-1,baseNum,newBasePhoneStr,newBaseHouseStr)
        else:
            return phRecursion(startNum,endNum-1,baseNum,basePhoneStr,baseHouseStr)



arr = [2]
for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,sheet_obj.max_row+2):
        if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
            break
    if i not in arr:
        arr.append(i)
    x=i

# print(arr)
print(len(arr))

newArr = numpy.array(arr)
for x in range(0,len(newArr)-1):
    if newArr[x+1]-newArr[x]==1:
        sheet_obj.cell(row=newArr[x],column=7).value = str(sheet_obj.cell(row=newArr[x],column=5).value)
        sheet_obj.cell(row=newArr[x],column=8).value = str(sheet_obj.cell(row=newArr[x],column=6).value)
    else:
        for i in range(newArr[x],newArr[x+1]):
            sheet_obj.cell(row=newArr[i],column=7).value = str(sheet_obj.cell(row=newArr[i],column=5).value)
            sheet_obj.cell(row=newArr[i],column=8).value = str(sheet_obj.cell(row=newArr[i],column=6).value)
            phRecursion(newArr[x],newArr[x+1]-1,i,str(sheet_obj.cell(row=i,column=7).value),str(sheet_obj.cell(row=i,column=8).value))



# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=9).value = str(sheet_obj.cell(row=x,column=7).value).count(';') + 1
#     sheet_obj.cell(row=x,column=10).value = str(sheet_obj.cell(row=x,column=8).value).count(';') + 1



wb_obj.save(path)

end = time.time()

print("Run time = " + time.strftime("%H:%M:%S", time.gmtime(end-start)))