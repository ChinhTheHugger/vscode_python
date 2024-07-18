import json
import openpyxl
import os
import time

# spreadsheet = openpyxl.Workbook()
# sheet = spreadsheet.active

json_list = [('ha_nam',2)]

for pair in json_list:
    name, count = pair[0], pair[1]
    
    parts = name.split('_')
    capitalized_parts = [part.capitalize() for part in parts]
    
    capitalized_name = '_'.join(capitalized_parts)
    
    new_folder_path = f'C:\\Users\\phams\\Downloads\\{capitalized_name}'
    os.makedirs(new_folder_path, exist_ok=True)
    
    new_folder_path_source = f'C:\\Users\\phams\\Downloads\\{capitalized_name}\\page sources'
    os.makedirs(new_folder_path_source, exist_ok=True)
    
    new_folder_path_info = f'C:\\Users\\phams\\Downloads\\{capitalized_name}\\thong tin du an'
    os.makedirs(new_folder_path_info, exist_ok=True)
    
    # spreadsheet = openpyxl.load_workbook("C:\\Users\\phams\\Downloads\\Hung_Yen\\hung_yen_projects_links.xlsx")
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active
    
    links = []

    for i in range(count):
        file_path = f"C:\\Users\\phams\\Downloads\\{name}_page_{i+1}.json"

        # Attempt to open the file with 'utf-8' encoding
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
        except UnicodeDecodeError:
            # If 'utf-8' encoding fails, try 'iso-8859-1'
            with open(file_path, 'r', encoding='iso-8859-1') as file:
                data = json.load(file)

        if len(data['children']) != 0:
            for i in range(len(data['children'])):
                if len(data['children'][i]['children']) != 0:
                    if data['children'][i]['children'][0]['value'] not in links and data['children'][i]['children'][0]['value'] != '':
                        links.append(data['children'][i]['children'][0]['value'])

    # print(len(links))

    sheet.cell(row=1,column=1).value = 'Page source'
    sheet.cell(row=1,column=2).value = 'Links'
    sheet.cell(row=1,column=3).value = 'Project name'

    for x in range(len(links)):
        sheet.cell(row=x+2,column=1).value = x+1
        sheet.cell(row=x+2,column=2).value = links[x]
        
    sheet_file = f"C:\\Users\\phams\\Downloads\\{capitalized_name}\\{name}_projects_links.xlsx"
    spreadsheet.save(sheet_file)
    
    print(f'Finished extracting: {name}')

