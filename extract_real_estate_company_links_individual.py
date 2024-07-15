import asyncio
from playwright.async_api import async_playwright
import openpyxl
from bs4 import BeautifulSoup
import re
from docx import Document
import os
import time
import json

# types = [
#     ('chu-dau-tu',10),
#     ('nha-thau-tu-van',2),
#     ('tong-thau-du-an',4),
#     ('thau-co-dien-vat-tu-thi-cong',1),
#     ('thau-hang-muc-xay-dung-vat-tu-thi-cong',1),
#     ('dich-vu-logicstic',1),
#     ('cong-nghe-va-thiet-bi-san-xuat',1)
# ]

types = [('gian-hang',2)]

for obj in types:
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active

    sheet.cell(row=1,column=1).value = 'Page source'
    sheet.cell(row=1,column=2).value = 'Link'
    sheet.cell(row=1,column=3).value = 'Name'
    sheet.cell(row=1,column=4).value = 'Type'
    
    row = 2
    
    links = []

    for i in range(obj[1]):
        file_path = f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\{obj[0]}_{i+1}.json'
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
        except UnicodeDecodeError:
            # If 'utf-8' encoding fails, try 'iso-8859-1'
            with open(file_path, 'r', encoding='iso-8859-1') as file:
                data = json.load(file)
                
        for idx, link in enumerate(data['children']):
            if 'value' in link:
                if 'thong-tin-doanh-nghiep' in link['value'] and link['value'] not in links:
                    sheet.cell(row=row,column=1).value = row-1
                    sheet.cell(row=row,column=2).value = link['value']
                    sheet.cell(row=row,column=4).value = obj[0]
                    
                    links.append(link['value'])
                    
                    print(f'Saved link of {obj[0]} at index {idx}')
                    
                    row += 1
        
    # print(len(links))
    
    spreadsheet.save(f'C:\\Users\\phams\\Downloads\\doanh_nghiep\\{obj[0]}.xlsx')
        