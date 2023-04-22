import openpyxl
from collections import Counter
import time
import numpy
import igraph
import sys

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.04.21 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V24 - for processing - Copy.XLSX"
pathTmp = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\area measurement.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

wb_tmp = openpyxl.load_workbook(pathTmp)
sheet_tmp = wb_tmp.active

start = time.time()
print("Start time = " + time.strftime("%H:%M:%S", time.gmtime(start)))
time.sleep(5)

# ***

# * Remove duplicate to get list of unique 

# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,x+5000):
#         if sheet_obj.cell(row=i,column=3).value != sheet_obj.cell(row=x,column=3).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=4).value == sheet_obj.cell(row=x,column=4).value:
#                 sheet_obj.cell(row=i,column=2).value = None
#             else:
#                 break
#     x=i

# ***

# for x in range(2,sheet_obj.max_row+1):
#     bArr = str(sheet_obj.cell(row=x,column=16).value).split(';')
#     tmpArr = [ele for ele in bArr if str(ele) != "1"]
#     sheet_obj.cell(row=x,column=16).value = ';'.join(tmpArr)

# for x in range(2,sheet_obj.max_row+1):
#     if sheet_obj.cell(row=x,column=17).value != None:
#         sheet_obj.cell(row=x,column=18).value = str(sheet_obj.cell(row=x,column=16).value)+";"+str(sheet_obj.cell(row=x,column=17).value)

for x in range(2,sheet_obj.max_row+1):
    for i in range(2,sheet_tmp.max_row+1):
        if sheet_tmp.cell(row=i,column=2).value == sheet_obj.cell(row=x,column=3).value:
            sheet_obj.cell(row=x,column=4).value = sheet_tmp.cell(row=i,column=3).value



wb_obj.save(path)

end = time.time()

print("Run time = " + time.strftime("%H:%M:%S", time.gmtime(end-start)))
print("End time = " + time.strftime("%H:%M:%S", time.gmtime(end)))