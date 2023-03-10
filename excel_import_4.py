import openpyxl
from collections import Counter

path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.08 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V21 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# ***

# * Count houses

# for x in range(2,sheet_obj.max_row+1):
#     houseCount = sheet_obj.cell(row = x, column = 8).value
#     sheet_obj.cell(row = x, column = 13).value = houseCount.count(';')+1

# ***

# * Group house and people

# for x in range(2,sheet_obj.max_row+1):
#     if sheet_obj.cell(row = x, column = 13).value > 400:
#         sheet_obj.cell(row = x, column = 11).value = sheet_obj.cell(row = x, column = 2).value + ";" + str(sheet_obj.cell(row = x, column = 13).value) + " căn"
#     else:
#         sheet_obj.cell(row = x, column = 11).value = sheet_obj.cell(row = x, column = 2).value + ";" + sheet_obj.cell(row = x, column = 10).value

# ***

# * Group people sharing phone number

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row = x, column = 11).number_format = '@'
#     sheet_obj.cell(row = x, column = 11).value = sheet_obj.cell(row = x, column = 10).value
#     for i in range(x+1,x+5000):
#         if sheet_obj.cell(row = i, column = 4).value == sheet_obj.cell(row = x, column = 4).value:
#             sheet_obj.cell(row = x, column = 11).value = sheet_obj.cell(row = x, column = 11).value + "&" + sheet_obj.cell(row = i, column = 10).value
#             sheet_obj.cell(row = i, column = 3).value = None
#             continue
#         else:
#             break
#     x = i

# ***

# * Remove duplicate and sort people

for x in range(2,sheet_obj.max_row+1):
    if sheet_obj.cell(row=x,column=3).value != None:
        for i in range(x+1,x+5500):
            if sheet_obj.cell(row=i,column=3).value == None:
                sheet_obj.cell(row=i,column=11).value = sheet_obj.cell(row=x,column=11).value
    x = i

for x in range(2,sheet_obj.max_row+1):
    peopleString = sheet_obj.cell(row = x, column = 11).value
    peopleArrayToSort = peopleString.split('&')
    peopleArrayToSort = list(dict.fromkeys(peopleArrayToSort))
    peopleListToStr = '&'.join(map(str,peopleArrayToSort))
    sheet_obj.cell(row = x, column = 11 ).value = peopleListToStr

wb_obj.save(path)