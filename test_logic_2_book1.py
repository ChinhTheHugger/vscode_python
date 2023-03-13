# 

import openpyxl
from collections import Counter
import time
import numpy

pathTest = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\Book1.XLSX"
path = "E:\\Pham Thanh Quyet - 23.12.2022\\DSKH 22.12.23\\VRS VRH\\23.03.11 Riverside+ Harmony Full - Tổng hợp khách hàng và căn V22 - for processing.XLSX"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = time.time()

# Phone group  > house step 1

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=4).value = sheet_obj.cell(row=x,column=2).value
#     for i in range(x+1,sheet_obj.max_row+1):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=3).value != sheet_obj.cell(row=x,column=3).value:
#                 break
#             else:
#                 temp = sheet_obj.cell(row=x,column=4).value + ";" + sheet_obj.cell(row=i,column=2).value
#                 sheet_obj.cell(row=x,column=4).value = temp
#     x=i

# for x in range(2,sheet_obj.max_row+1):
#     peopleString = str(sheet_obj.cell(row = x, column = 4).value)
#     peopleArrayToSort = peopleString.split(';')
#     peopleArrayToSort = list(dict.fromkeys(peopleArrayToSort))
#     peopleListToStr = ';'.join(map(str,peopleArrayToSort))
#     sheet_obj.cell(row = x, column = 4).value = peopleListToStr

# House group > phone step 2

# for x in range(2,sheet_obj.max_row+1):
#     sheet_obj.cell(row=x,column=5).value = sheet_obj.cell(row=x,column=3).value
#     for i in range(x+1,sheet_obj.max_row+1):
#         if sheet_obj.cell(row=i,column=1).value != sheet_obj.cell(row=x,column=1).value:
#             break
#         else:
#             if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#                 break
#             else:
#                 temp = sheet_obj.cell(row=x,column=5).value + ";" + sheet_obj.cell(row=i,column=3).value
#                 sheet_obj.cell(row=x,column=5).value = temp
#     x=i

# for x in range(2,sheet_obj.max_row+1):
#     peopleString = str(sheet_obj.cell(row = x, column = 5).value)
#     peopleArrayToSort = peopleString.split(';')
#     peopleArrayToSort = list(dict.fromkeys(peopleArrayToSort))
#     peopleListToStr = ';'.join(map(str,peopleArrayToSort))
#     sheet_obj.cell(row = x, column = 5).value = peopleListToStr

# ***

# House group

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=5).value=sheet_obj.cell(row=x,column=4).value
#     else:
#         for a in range(x,i+1):
#             houseArrTemp=[]
#             houseArrTemp.append(sheet_obj.cell(row=a,column=4).value)
#             for b in range(x,i+1):
#                 if sheet_obj.cell(row=b,column=3).value==sheet_obj.cell(row=a,column=3).value and sheet_obj.cell(row=b,column=4).value not in houseArrTemp:
#                     houseArrTemp.append(sheet_obj.cell(row=b,column=4).value)
#             # houseArr=list(dict.fromkeys(houseArrTemp))
#             houseStr=';'.join(map(str,houseArrTemp))
#             sheet_obj.cell(row=a,column=5).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)
# print(len(arr))

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-2):
#     houseGroup(newArr[x],newArr[x+1]-1)

# Phone group

def phoneGroup(x,i):
    if i-x==0:
        sheet_obj.cell(row=x,column=6).value=sheet_obj.cell(row=x,column=3).value
    else:
        for a in range(x,i+1):
            houseArrTemp=[]
            houseArrTemp.append(sheet_obj.cell(row=a,column=3).value)
            for b in range(x,i+1):
                if sheet_obj.cell(row=b,column=4).value==sheet_obj.cell(row=a,column=4).value and sheet_obj.cell(row=b,column=3).value not in houseArrTemp:
                    houseArrTemp.append(sheet_obj.cell(row=b,column=3).value)
            # houseArr=list(dict.fromkeys(houseArrTemp))
            houseStr=';'.join(map(str,houseArrTemp))
            sheet_obj.cell(row=a,column=6).value=houseStr
    return

arr = [2]
for x in range(2,sheet_obj.max_row+1):
    for i in range(x+1,sheet_obj.max_row+2):
        if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
            break
    if i not in arr:
        arr.append(i)
    x=i

print(arr)
print(len(arr))

newArr = numpy.array(arr)
for x in range(0,len(arr)-2):
    phoneGroup(newArr[x],newArr[x+1]-1)

# ***

# Clean cells - why doesnt this work anymore???

# for x in range(2,sheet_obj.max_row+1):
#     tempStr1=sheet_obj.cell(row=x,column=5).value
#     tempArr1=tempStr1.split(';')
#     tempArr1=list(dict.fromkeys(tempStr1))
#     tempCln1=';'.join(map(str,tempArr1))
#     sheet_obj.cell(row=x,column=5).value=tempCln1
#     tempStr2=sheet_obj.cell(row=x,column=6).value
#     tempArr2=tempStr1.split(';')
#     tempArr2=list(dict.fromkeys(tempStr2))
#     tempCln2=';'.join(map(str,tempArr2))
#     sheet_obj.cell(row=x,column=6).value=tempCln2

# ***

# House group 2
# this took way too long the first time running
# didnt even finished
# might be too much with the whole separating phone group each cell into array then run through each element
# new idea: split the phone group into rows with power query, then run "House group" above again
#
# seem to work, dont know about accuracy tho

# def houseGroup(x,i):
#     if i-x==0:
#         sheet_obj.cell(row=x,column=7).value=sheet_obj.cell(row=x,column=4).value
#     else:
#         for a in range(x,i+1):
#             houseArrTemp=[]
#             houseArrTemp.append(sheet_obj.cell(row=a,column=4).value)
#             for b in range(x,i+1):
#                 if sheet_obj.cell(row=b,column=3).value==sheet_obj.cell(row=a,column=3).value:
#                     houseArrTemp.append(sheet_obj.cell(row=b,column=4).value)
#             houseArr=list(dict.fromkeys(houseArrTemp))
#             houseStr=';'.join(map(str,houseArr))
#             sheet_obj.cell(row=a,column=7).value=houseStr
#     return

# arr = [2]
# for x in range(2,sheet_obj.max_row+1):
#     for i in range(x+1,sheet_obj.max_row+2):
#         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
#             break
#     if i not in arr:
#         arr.append(i)
#     x=i

# print(arr)

# newArr = numpy.array(arr)
# for x in range(0,len(arr)-2):
#     houseGroup(newArr[x],newArr[x+1]-1)

# the first method - for now, its deemed a failure
    #

    # def houseGroup(x,i):
    #     if i-x==0:
    #         sheet_obj.cell(row=x,column=7).value=sheet_obj.cell(row=x,column=4).value
    #     else:
    #         for pos in range(x,i+1):
    #             phoneStr=sheet_obj.cell(row=pos,column=6).value
    #             phoneArr=phoneStr.split(';')
    #             houseArrTemp=[]
    #             for ele in phoneArr:
    #                 for a in range(x,i+1):
    #                     houseArrTemp.append(sheet_obj.cell(row=a,column=4).value)
    #                     for b in range(x,i+1):
    #                         if sheet_obj.cell(row=b,column=3).value==ele:
    #                             houseArrTemp.append(sheet_obj.cell(row=b,column=4).value)
    #             houseArr=list(dict.fromkeys(houseArrTemp))
    #             houseStr=';'.join(map(str,houseArr))
    #             sheet_obj.cell(row=pos,column=7).value=houseStr
    #     return

    # arr = [2]
    # for x in range(2,sheet_obj.max_row+1):
    #     for i in range(x+1,sheet_obj.max_row+2):
    #         if sheet_obj.cell(row=i,column=2).value != sheet_obj.cell(row=x,column=2).value:
    #             break
    #     if i not in arr:
    #         arr.append(i)
    #     x=i

    # print(arr)

    # newArr = numpy.array(arr)
    # for x in range(0,len(arr)-2):
    #     houseGroup(newArr[x],newArr[x+1]-1)

    #

wb_obj.save(path)

end = time.time()

print(time.strftime("%H:%M:%S", time.gmtime(end-start)))